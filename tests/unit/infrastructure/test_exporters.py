"""Tests unitarios para NullExporter, OpenpyxlExporter y crear_exporter."""
from __future__ import annotations

import io
import csv
import tempfile
from pathlib import Path

import pytest
import openpyxl

from src.domain.ports.service_ports import IExporterService
from src.infrastructure.exporters.null_exporter import NullExporter
from src.infrastructure.exporters.openpyxl_exporter import OpenpyxlExporter
from src.infrastructure.exporters.exporter_factory import crear_exporter


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

DATOS_SIMPLES = [
    {"nombre": "Ana García", "nota": 4.5, "ausencias": 2},
    {"nombre": "Luis Pérez", "nota": 3.8, "ausencias": 0},
    {"nombre": "María López", "nota": 5.0, "ausencias": 1},
]


# ===========================================================================
# NullExporter
# ===========================================================================

class TestNullExporter:
    def test_implementa_el_port(self):
        assert isinstance(NullExporter(), IExporterService)

    # --- excel ---

    def test_excel_lanza_runtime_error(self):
        with pytest.raises(RuntimeError, match="openpyxl"):
            NullExporter().exportar_excel(DATOS_SIMPLES)

    def test_excel_mensaje_indica_como_instalar(self):
        with pytest.raises(RuntimeError) as exc:
            NullExporter().exportar_excel([])
        assert "pip install openpyxl" in str(exc.value)

    # --- pdf ---

    def test_pdf_lanza_runtime_error(self):
        with pytest.raises(RuntimeError, match="weasyprint|reportlab"):
            NullExporter().exportar_pdf("<p>test</p>")

    # --- csv ---

    def test_csv_retorna_bytes(self):
        resultado = NullExporter().exportar_csv(DATOS_SIMPLES)
        assert isinstance(resultado, bytes)

    def test_csv_vacio_retorna_bytes_vacios(self):
        assert NullExporter().exportar_csv([]) == b""

    def test_csv_contiene_headers(self):
        resultado = NullExporter().exportar_csv(DATOS_SIMPLES)
        texto = resultado.decode("utf-8-sig")
        assert "nombre" in texto
        assert "nota" in texto

    def test_csv_contiene_datos(self):
        resultado = NullExporter().exportar_csv(DATOS_SIMPLES)
        texto = resultado.decode("utf-8-sig")
        assert "Ana García" in texto
        assert "Luis Pérez" in texto

    def test_csv_encoding_por_defecto_utf8_bom(self):
        resultado = NullExporter().exportar_csv(DATOS_SIMPLES)
        # utf-8-sig comienza con BOM (EF BB BF)
        assert resultado[:3] == b"\xef\xbb\xbf"

    def test_csv_encoding_alternativo(self):
        resultado = NullExporter().exportar_csv(DATOS_SIMPLES, encoding="utf-8")
        assert resultado[:3] != b"\xef\xbb\xbf"

    def test_csv_escribe_archivo_si_ruta_destino(self, tmp_path):
        ruta = tmp_path / "salida.csv"
        resultado = NullExporter().exportar_csv(DATOS_SIMPLES, ruta_destino=ruta)
        assert resultado == b""
        assert ruta.exists()
        assert ruta.stat().st_size > 0

    def test_csv_parseable(self):
        resultado = NullExporter().exportar_csv(DATOS_SIMPLES)
        texto = resultado.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(texto))
        filas = list(reader)
        assert len(filas) == 3
        assert filas[0]["nombre"] == "Ana García"


# ===========================================================================
# OpenpyxlExporter — Excel
# ===========================================================================

class TestOpenpyxlExporterExcel:
    def test_implementa_el_port(self):
        assert isinstance(OpenpyxlExporter(), IExporterService)

    def test_retorna_bytes_no_vacios(self):
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES)
        assert isinstance(resultado, bytes) and len(resultado) > 0

    def test_bytes_son_xlsx_valido(self):
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES)
        wb = openpyxl.load_workbook(io.BytesIO(resultado))
        assert wb is not None

    def test_nombre_hoja_correcto(self):
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES, nombre_hoja="Notas")
        wb = openpyxl.load_workbook(io.BytesIO(resultado))
        assert "Notas" in wb.sheetnames

    def test_nombre_hoja_se_trunca_a_31_chars(self):
        nombre_largo = "A" * 50
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES, nombre_hoja=nombre_largo)
        wb = openpyxl.load_workbook(io.BytesIO(resultado))
        assert len(wb.active.title) <= 31

    def test_headers_en_primera_fila(self):
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES)
        ws = openpyxl.load_workbook(io.BytesIO(resultado)).active
        assert ws.cell(1, 1).value == "nombre"
        assert ws.cell(1, 2).value == "nota"
        assert ws.cell(1, 3).value == "ausencias"

    def test_datos_desde_fila_dos(self):
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES)
        ws = openpyxl.load_workbook(io.BytesIO(resultado)).active
        assert ws.cell(2, 1).value == "Ana García"
        assert ws.cell(2, 2).value == pytest.approx(4.5)

    def test_total_filas_correcto(self):
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES)
        ws = openpyxl.load_workbook(io.BytesIO(resultado)).active
        # 1 header + 3 datos
        assert ws.max_row == 4

    def test_datos_vacios_escribe_sin_datos(self):
        resultado = OpenpyxlExporter().exportar_excel([])
        ws = openpyxl.load_workbook(io.BytesIO(resultado)).active
        assert ws.cell(1, 1).value == "Sin datos"

    def test_escribe_archivo_si_ruta_destino(self, tmp_path):
        ruta = tmp_path / "informe.xlsx"
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES, ruta_destino=ruta)
        assert resultado == b""
        assert ruta.exists()
        wb = openpyxl.load_workbook(ruta)
        assert wb.active.max_row == 4

    def test_headers_tienen_fill_azul(self):
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES)
        ws = openpyxl.load_workbook(io.BytesIO(resultado)).active
        fill = ws.cell(1, 1).fill
        assert fill.fgColor.rgb.endswith("2B6CB0")

    def test_headers_son_negrita(self):
        resultado = OpenpyxlExporter().exportar_excel(DATOS_SIMPLES)
        ws = openpyxl.load_workbook(io.BytesIO(resultado)).active
        assert ws.cell(1, 1).font.bold is True

    def test_dataset_grande_sin_error(self):
        datos_grandes = [{"col_a": i, "col_b": f"valor_{i}"} for i in range(500)]
        resultado = OpenpyxlExporter().exportar_excel(datos_grandes)
        ws = openpyxl.load_workbook(io.BytesIO(resultado)).active
        assert ws.max_row == 501  # 1 header + 500 filas


# ===========================================================================
# OpenpyxlExporter — PDF
# ===========================================================================

class TestOpenpyxlExporterPdf:
    def test_pdf_lanza_not_implemented(self):
        with pytest.raises(NotImplementedError):
            OpenpyxlExporter().exportar_pdf("<p>hola</p>")

    def test_pdf_mensaje_sugiere_alternativa(self):
        with pytest.raises(NotImplementedError) as exc:
            OpenpyxlExporter().exportar_pdf("")
        assert "weasyprint" in str(exc.value).lower() or "reportlab" in str(exc.value).lower()


# ===========================================================================
# OpenpyxlExporter — CSV
# ===========================================================================

class TestOpenpyxlExporterCsv:
    def test_csv_retorna_bytes(self):
        assert isinstance(OpenpyxlExporter().exportar_csv(DATOS_SIMPLES), bytes)

    def test_csv_vacio(self):
        assert OpenpyxlExporter().exportar_csv([]) == b""

    def test_csv_contiene_datos(self):
        texto = OpenpyxlExporter().exportar_csv(DATOS_SIMPLES).decode("utf-8-sig")
        assert "María López" in texto

    def test_csv_escribe_archivo(self, tmp_path):
        ruta = tmp_path / "out.csv"
        resultado = OpenpyxlExporter().exportar_csv(DATOS_SIMPLES, ruta_destino=ruta)
        assert resultado == b""
        assert ruta.exists()


# ===========================================================================
# crear_exporter (factory)
# ===========================================================================

class TestCrearExporter:
    def test_retorna_implementacion_del_port(self):
        exporter = crear_exporter()
        assert isinstance(exporter, IExporterService)

    def test_retorna_openpyxl_cuando_disponible(self):
        # openpyxl está instalado en el entorno de prueba
        exporter = crear_exporter()
        assert isinstance(exporter, OpenpyxlExporter)

    def test_factory_puede_exportar_csv(self):
        exporter = crear_exporter()
        resultado = exporter.exportar_csv(DATOS_SIMPLES)
        assert isinstance(resultado, bytes) and len(resultado) > 0

    def test_factory_puede_exportar_excel(self):
        exporter = crear_exporter()
        resultado = exporter.exportar_excel(DATOS_SIMPLES)
        wb = openpyxl.load_workbook(io.BytesIO(resultado))
        assert wb is not None

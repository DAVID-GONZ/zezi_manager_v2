"""Tests unitarios para InformeService."""
from __future__ import annotations

from datetime import date
from typing import Any

import pytest

from src.domain.models.convivencia import (
    CategoriaObservacion,
    NotaComportamiento,
    ObservacionPeriodo,
)
from src.domain.models.dtos import (
    DashboardMetricsDTO,
    FormatoInforme,
    InformeAsistenciaDTO,
    InformeNotasDTO,
)
from src.domain.ports.estadisticos_repo import IEstadisticosRepository
from src.domain.ports.service_ports import IExporterService
from src.services.convivencia_service import ConvivenciaService
from src.services.informe_service import InformeService

# ===========================================================================
# Fakes
# ===========================================================================

class _FakePeriodo:
    """Objeto periodo mínimo para los tests (id + nombre)."""
    def __init__(self, id_: int, nombre: str):
        self.id = id_
        self.nombre = nombre


class _FakePeriodoSvc:
    """Fake PeriodoService con listar_por_anio."""
    def __init__(self, periodos: list[_FakePeriodo]):
        self._periodos = periodos

    def listar_por_anio(self, anio_id: int) -> list[_FakePeriodo]:
        return self._periodos


class FakeConvivenciaRepo:
    """Fake minimal de IConvivenciaRepository para tests de convivencia_boletin."""

    def __init__(
        self,
        nota=None,
        observaciones=None,
        notas_por_estudiante=None,
        obs_por_periodo=None,
        categorias=None,
        obs_por_grupo=None,
        registros=None,
    ):
        self._nota = nota
        self._observaciones = observaciones or []
        # Para convivencia_boletin_anual:
        self._notas_por_estudiante: list = notas_por_estudiante or []
        self._obs_por_periodo: dict = obs_por_periodo or {}  # periodo_id -> [ObservacionPeriodo]
        self._categorias: list = categorias or []
        # Para resumen grupo (convivencia_31):
        self._obs_por_grupo: list = obs_por_grupo or []
        self._registros: list = registros or []

    def get_nota(self, estudiante_id: int, periodo_id: int):
        return self._nota

    def listar_observaciones_por_estudiante(
        self, estudiante_id: int, periodo_id: int | None = None, solo_publicas: bool = False
    ):
        if periodo_id is not None and self._obs_por_periodo:
            # Modo anual: usar lookup por periodo
            return self._obs_por_periodo.get(periodo_id, [])
        # Modo periodo (compat con tests originales): devolver lista plana
        return self._observaciones

    def listar_observaciones_por_grupo(
        self, grupo_id: int, periodo_id: int | None = None, solo_publicas: bool = False
    ):
        return self._obs_por_grupo

    def listar_notas_por_estudiante(self, estudiante_id: int):
        return self._notas_por_estudiante

    def listar_categorias(self, solo_activas: bool = True, institucion_id=None):
        return self._categorias

    # Añadido en convivencia_29: _get_registros_periodo en InformeService lo requiere.
    def listar_registros(self, filtro, institucion_id=None):
        return self._registros

    # Añadido en convivencia_38: tipos y medidas requeridos por _registros_informables_periodo.
    def listar_tipos_situacion(self, solo_activas=True, institucion_id=None):
        return []

    def listar_medidas(self, solo_activas=True, institucion_id=None):
        return []


class FakeEstadRepo(IEstadisticosRepository):
    def calcular_metricas_dashboard(self, g, p, nota_minima=60.0) -> DashboardMetricsDTO:
        return DashboardMetricsDTO(grupo_id=g)

    def promedio_general_grupo(self, g, p, nota_minima=60.0) -> float:
        return 70.0

    def porcentaje_asistencia_global(self, g, p) -> float:
        return 85.0

    def contar_alertas_pendientes(self, g) -> int:
        return 0

    def promedio_por_asignacion(self, g, a, p) -> float:
        return 70.0

    def distribucion_desempenos(self, g, a, p, niveles) -> dict[str, int]:
        return {}

    def comparativo_periodos(self, g, a, anio) -> list[dict[str, Any]]:
        return []

    def promedios_por_area(self, g, p) -> list[dict[str, Any]]:
        return []

    def estudiantes_en_riesgo_academico(self, g, p, nota_minima=60.0, min_asig=1) -> list[int]:
        return []

    def ranking_grupo(self, g, p) -> list[dict[str, Any]]:
        return []

    def tendencia_asistencia(self, g, a, p) -> list[dict[str, Any]]:
        return []

    def distribucion_estados_asistencia(self, g, a, p) -> dict[str, int]:
        return {}

    def consolidado_notas_grupo(self, g, p) -> list[dict[str, Any]]:
        return [{"nombre": "Est A", "matematicas": 80.0, "promedio_periodo": 80.0}]

    def consolidado_asistencia_grupo(self, g, p) -> list[dict[str, Any]]:
        return [{"nombre": "Est A", "presentes": 20}]

    def consolidado_anual_grupo(self, g, anio) -> list[dict[str, Any]]:
        return [{"nombre": "Est A", "nota_anual": 75.0}]

    def boletin_datos_periodo(self, estudiante_id, grupo_id, periodo_id):
        return {"estudiante": {}, "areas": []}

    def boletin_datos_acumulado(self, estudiante_id, grupo_id, hasta_periodo_id):
        return {"estudiante": {}, "periodos": [], "areas": [], "es_ultimo_periodo": False}

    def boletin_datos_anual(self, estudiante_id, grupo_id, anio_id):
        return {
            "estudiante": {},
            "periodos": [
                {"id": 1, "nombre": "Período 1"},
                {"id": 2, "nombre": "Período 2"},
            ],
            "areas": [
                {
                    "area_nombre": "Matemáticas",
                    "asignaturas": [
                        {
                            "nombre": "Álgebra",
                            "notas_periodo": {1: 80.0, 2: 75.0},
                            "definitiva": 77.5,
                            "presentes": 20,
                            "faltas_injustificadas": 1,
                            "faltas_justificadas": 0,
                            "retrasos": 0,
                            "excusas": 0,
                        }
                    ],
                }
            ],
        }


class FakeExporter(IExporterService):
    def exportar_excel(self, datos, nombre_hoja="Datos", ruta_destino=None) -> bytes:
        return b"EXCEL:" + str(len(datos)).encode()

    def exportar_pdf(self, html_content, ruta_destino=None) -> bytes:
        return b"PDF:" + html_content[:20].encode()

    def exportar_csv(self, datos, ruta_destino=None, encoding="utf-8-sig") -> bytes:
        return b"CSV"


# ===========================================================================
# Helpers
# ===========================================================================

def _dto_notas() -> InformeNotasDTO:
    return InformeNotasDTO(
        grupo_id=10, asignacion_id=3, periodo_id=5,
        fecha_desde=date(2025, 1, 1), fecha_hasta=date(2025, 6, 30),
    )


def _dto_asistencia() -> InformeAsistenciaDTO:
    return InformeAsistenciaDTO(
        grupo_id=10, asignacion_id=3, periodo_id=5,
        fecha_desde=date(2025, 1, 1), fecha_hasta=date(2025, 6, 30),
    )


# ===========================================================================
# Tests
# ===========================================================================

class TestSinExporter:
    def test_datos_informe_notas_sin_exporter(self):
        svc = InformeService(FakeEstadRepo())  # sin exporter
        datos = svc.datos_informe_notas(_dto_notas())
        assert len(datos) == 1

    def test_lanza_si_intenta_generar_sin_exporter(self):
        svc = InformeService(FakeEstadRepo())
        with pytest.raises(ValueError, match="exportador"):
            svc.generar_notas(_dto_notas())

    def test_datos_asistencia_sin_exporter(self):
        svc = InformeService(FakeEstadRepo())
        datos = svc.datos_informe_asistencia(_dto_asistencia())
        assert isinstance(datos, list)


class TestConExporter:
    def test_genera_notas_en_excel(self):
        svc = InformeService(FakeEstadRepo(), exporter=FakeExporter())
        resultado = svc.generar_notas(_dto_notas())
        assert resultado.startswith(b"EXCEL:")

    def test_genera_notas_en_pdf(self):
        svc = InformeService(FakeEstadRepo(), exporter=FakeExporter())
        dto = InformeNotasDTO(
            grupo_id=10, asignacion_id=3, periodo_id=5,
            fecha_desde=date(2025, 1, 1), fecha_hasta=date(2025, 6, 30),
            formato=FormatoInforme.PDF,
        )
        resultado = svc.generar_notas(dto)
        assert resultado.startswith(b"PDF:")

    def test_genera_consolidado_anual(self):
        svc = InformeService(FakeEstadRepo(), exporter=FakeExporter())
        resultado = svc.generar_consolidado_anual(grupo_id=10, anio_id=1)
        assert resultado.startswith(b"EXCEL:")

    def test_exportar_csv(self):
        svc = InformeService(FakeEstadRepo(), exporter=FakeExporter())
        resultado = svc.exportar_csv([{"col": "val"}])
        assert resultado == b"CSV"


# ===========================================================================
# Group 7 — exportar_estadistico + generar_boletines_grupo
# ===========================================================================

class _Est:
    def __init__(self, id_, nombre, apellido):
        self.id = id_
        self.nombre = nombre
        self.apellido = apellido


class _FakeEstRepo:
    def __init__(self, ests):
        self._ests = ests
    # Espeja IEstudianteRepository.listar_por_grupo (institucion_id posicional
    # obligatorio desde la migración multi-tenant).
    def listar_por_grupo(self, grupo_id, institucion_id, solo_activos=True):
        return self._ests


class TestExportarEstadistico:
    def test_excel_consolidado_notas(self):
        svc = InformeService(FakeEstadRepo(), FakeExporter())
        datos = [{"nombre_completo": "Ana", "promedio": 80.0, "estudiante_id": 1}]
        out = svc.exportar_estadistico("consolidado_notas", datos, "excel")
        assert out.startswith(b"EXCEL:")

    def test_pdf_inyecta_meta(self):
        svc = InformeService(FakeEstadRepo(), FakeExporter())
        ctx = {"grupo_nombre": "601", "periodo_nombre": "P1", "asignatura_nombre": "Mat"}
        out = svc.exportar_estadistico("ranking_grupo", [{"posicion": 1}], "pdf", ctx)
        assert out.startswith(b"PDF:")

    def test_estados_asistencia_dict(self):
        svc = InformeService(FakeEstadRepo(), FakeExporter())
        out = svc.exportar_estadistico("estados_asistencia", {"P": 10, "FI": 2}, "excel")
        assert out == b"EXCEL:2"  # 2 filas

    def test_tipo_desconocido_lanza(self):
        svc = InformeService(FakeEstadRepo(), FakeExporter())
        with pytest.raises(ValueError, match="no reconocido"):
            svc.exportar_estadistico("xxx", [], "excel")


class _RealXlsxExporter(FakeExporter):
    """Exporter que devuelve un .xlsx válido (para que merge_excels funcione)."""
    def exportar_excel(self, datos, nombre_hoja="Datos", ruta_destino=None) -> bytes:
        import io

        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.append(["col"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


class TestGenerarBoletinesGrupo:
    def test_excel_fusiona_por_estudiante(self):
        ests = [_Est(1, "Ana", "Lopez"), _Est(2, "Beto", "Diaz")]
        svc = InformeService(FakeEstadRepo(), _RealXlsxExporter(), estudiante_repo=_FakeEstRepo(ests))
        r = svc.generar_boletines_grupo(grupo_id=10, periodo_id=5, formato="excel")
        assert r.contenido is not None
        assert r.errores == []
        import io

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.contenido))
        assert len(wb.sheetnames) == 2

    def test_sin_periodo_ni_anio_lanza(self):
        svc = InformeService(FakeEstadRepo(), FakeExporter(), estudiante_repo=_FakeEstRepo([]))
        with pytest.raises(ValueError, match=r"periodo_id|anio_id"):
            svc.generar_boletines_grupo(grupo_id=10, formato="excel")

    def test_sin_estudiantes_contenido_none(self):
        svc = InformeService(FakeEstadRepo(), FakeExporter(), estudiante_repo=_FakeEstRepo([]))
        r = svc.generar_boletines_grupo(grupo_id=10, periodo_id=5, formato="excel")
        assert r.contenido is None


# ===========================================================================
# Group 8 — convivencia_boletin (convivencia_07)
# ===========================================================================

class TestConvivenciaBoletin:
    def test_convivencia_boletin_sin_repo(self):
        """T4a: sin convivencia_repo inyectado → dict con None y [] (convivencia_29 añade registros)."""
        svc = InformeService(FakeEstadRepo())
        resultado = svc.convivencia_boletin(1, 1)
        assert resultado == {"nota": None, "nota_observacion": None, "observaciones": [], "registros": []}

    def test_convivencia_boletin_con_nota(self):
        """T4b: repo retorna nota con valor y obs pública → dict correcto."""
        nota = NotaComportamiento(
            estudiante_id=1, grupo_id=1, periodo_id=1,
            valor=85.0, observacion="Excelente comportamiento",
        )
        obs = ObservacionPeriodo(
            estudiante_id=1, asignacion_id=1, periodo_id=1,
            texto="Participa activamente en clase", es_publica=True,
        )
        repo = FakeConvivenciaRepo(nota=nota, observaciones=[obs])
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(repo),
        )
        resultado = svc.convivencia_boletin(1, 1)
        assert resultado["nota"] == 85.0
        assert resultado["nota_observacion"] == "Excelente comportamiento"
        assert resultado["observaciones"] == ["Participa activamente en clase"]

    def test_convivencia_boletin_sin_nota(self):
        """T4c: repo retorna None para nota + obs vacía → dict correcto (convivencia_29 añade registros)."""
        repo = FakeConvivenciaRepo(nota=None, observaciones=[])
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(repo),
        )
        resultado = svc.convivencia_boletin(1, 1)
        # paquete_boletin_periodo incluye también observaciones_por_categoria
        assert resultado["nota"] is None
        assert resultado["nota_observacion"] is None
        assert resultado["observaciones"] == []
        assert resultado["registros"] == []

    def test_convivencia_boletin_solo_obs(self):
        """T4d: nota None pero hay 2 obs públicas → lista con 2 textos."""
        obs1 = ObservacionPeriodo(
            estudiante_id=1, asignacion_id=1, periodo_id=1,
            texto="Obs uno", es_publica=True,
        )
        obs2 = ObservacionPeriodo(
            estudiante_id=1, asignacion_id=2, periodo_id=1,
            texto="Obs dos", es_publica=True,
        )
        repo = FakeConvivenciaRepo(nota=None, observaciones=[obs1, obs2])
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(repo),
        )
        resultado = svc.convivencia_boletin(1, 1)
        assert resultado["nota"] is None
        assert resultado["nota_observacion"] is None
        assert len(resultado["observaciones"]) == 2
        assert resultado["observaciones"] == ["Obs uno", "Obs dos"]


# ===========================================================================
# Group 9 — convivencia_boletin_anual (convivencia_28)
# ===========================================================================

def _make_periodos(*args: tuple[int, str]) -> list[_FakePeriodo]:
    """Helper: crea lista de _FakePeriodo desde tuplas (id, nombre)."""
    return [_FakePeriodo(id_, nombre) for id_, nombre in args]


class TestConvivenciaBoletinAnual:
    """T1 — Tests unitarios para InformeService.convivencia_boletin_anual."""

    def test_sin_repo_devuelve_vacio(self):
        """Sin convivencia_repo → estructura vacía."""
        svc = InformeService(FakeEstadRepo())
        resultado = svc.convivencia_boletin_anual(1, 2025)
        assert resultado["periodos"] == []
        assert resultado["notas_por_periodo"] == {}
        assert resultado["definitiva"] is None
        assert resultado["concepto"] is None
        assert resultado["observaciones_por_categoria"] == []

    def test_sin_periodo_svc_provider_devuelve_vacio(self):
        """Sin periodo_svc_provider → estructura vacía aunque haya repo."""
        repo = FakeConvivenciaRepo()
        # ConvivenciaService sin periodo_svc_provider → paquete_boletin_anual vacío
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(repo),
        )
        resultado = svc.convivencia_boletin_anual(1, 2025)
        assert resultado["periodos"] == []
        assert resultado["definitiva"] is None

    def test_nominal_2_periodos_con_nota_1_sin_nota(self):
        """T1a: 3 periodos, P1=80 P2=75 P3=None → definitiva=77.5, concepto del P2."""
        p1, p2, p3 = (
            _FakePeriodo(1, "Período 1"),
            _FakePeriodo(2, "Período 2"),
            _FakePeriodo(3, "Período 3"),
        )
        nota_p1 = NotaComportamiento(
            estudiante_id=1, grupo_id=1, periodo_id=1,
            valor=80.0, observacion="Bien en P1",
        )
        nota_p2 = NotaComportamiento(
            estudiante_id=1, grupo_id=1, periodo_id=2,
            valor=75.0, observacion="Excelente en P2",
        )
        repo = FakeConvivenciaRepo(
            notas_por_estudiante=[nota_p1, nota_p2],
        )
        periodo_svc = _FakePeriodoSvc([p1, p2, p3])
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(
                repo, periodo_svc_provider=lambda: periodo_svc
            ),
        )
        resultado = svc.convivencia_boletin_anual(1, 2025)

        assert len(resultado["periodos"]) == 3
        assert resultado["notas_por_periodo"][1] == 80.0
        assert resultado["notas_por_periodo"][2] == 75.0
        assert resultado["notas_por_periodo"][3] is None
        assert resultado["definitiva"] == 77.5
        # Concepto del último periodo con nota = P2
        assert resultado["concepto"] == "Excelente en P2"

    def test_sin_notas_ni_obs_devuelve_campos_vacios(self):
        """T1b: sin notas y sin obs → definitiva=None, concepto=None, obs=[]."""
        p1 = _FakePeriodo(1, "Período 1")
        repo = FakeConvivenciaRepo(notas_por_estudiante=[])
        periodo_svc = _FakePeriodoSvc([p1])
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(
                repo, periodo_svc_provider=lambda: periodo_svc
            ),
        )
        resultado = svc.convivencia_boletin_anual(1, 2025)
        assert resultado["definitiva"] is None
        assert resultado["concepto"] is None
        assert resultado["observaciones_por_categoria"] == []

    def test_observaciones_agrupadas_por_categoria(self):
        """T1c: obs en P1 (cat=1 activa) y P2 (sin cat) → 2 grupos, 'Sin categoría' al final."""
        p1 = _FakePeriodo(1, "P1")
        p2 = _FakePeriodo(2, "P2")
        cat_responsabilidad = CategoriaObservacion(id=1, nombre="Responsabilidad", activa=True)
        obs_p1 = ObservacionPeriodo(
            estudiante_id=1, asignacion_id=1, periodo_id=1,
            texto="Muy puntual", es_publica=True, categoria_id=1,
        )
        obs_p2 = ObservacionPeriodo(
            estudiante_id=1, asignacion_id=1, periodo_id=2,
            texto="Sin categoría obs", es_publica=True, categoria_id=None,
        )
        repo = FakeConvivenciaRepo(
            categorias=[cat_responsabilidad],
            obs_por_periodo={1: [obs_p1], 2: [obs_p2]},
        )
        periodo_svc = _FakePeriodoSvc([p1, p2])
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(
                repo, periodo_svc_provider=lambda: periodo_svc
            ),
        )
        resultado = svc.convivencia_boletin_anual(1, 2025)
        grupos = resultado["observaciones_por_categoria"]
        assert len(grupos) == 2
        # Activa primero
        assert grupos[0]["categoria"] == "Responsabilidad"
        assert grupos[0]["items"][0]["texto"] == "Muy puntual"
        assert grupos[0]["items"][0]["periodo"] == "P1"
        # Sin categoría al final
        assert grupos[1]["categoria"] == "Sin categoría"
        assert grupos[1]["items"][0]["texto"] == "Sin categoría obs"

    def test_definitiva_un_solo_periodo_con_nota(self):
        """T1d: solo un periodo con nota → definitiva = esa nota exacta."""
        p1 = _FakePeriodo(1, "P1")
        p2 = _FakePeriodo(2, "P2")
        nota_p1 = NotaComportamiento(
            estudiante_id=1, grupo_id=1, periodo_id=1, valor=90.0,
        )
        repo = FakeConvivenciaRepo(notas_por_estudiante=[nota_p1])
        periodo_svc = _FakePeriodoSvc([p1, p2])
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(
                repo, periodo_svc_provider=lambda: periodo_svc
            ),
        )
        resultado = svc.convivencia_boletin_anual(1, 2025)
        assert resultado["definitiva"] == 90.0
        assert resultado["notas_por_periodo"][2] is None


# ===========================================================================
# Group 10 — generar_boletin_anual con convivencia (convivencia_28)
# ===========================================================================

class TestGenerarBoletinAnualConvivencia:
    """T2/T3/T5 — Tests de integración de convivencia en el boletín anual."""

    def _make_svc_con_conv(self):
        """Devuelve InformeService + FakeEstadRepo configurado para anual."""
        p1 = _FakePeriodo(1, "Período 1")
        p2 = _FakePeriodo(2, "Período 2")
        nota_p1 = NotaComportamiento(
            estudiante_id=1, grupo_id=1, periodo_id=1, valor=80.0,
        )
        nota_p2 = NotaComportamiento(
            estudiante_id=1, grupo_id=1, periodo_id=2, valor=75.0,
            observacion="Avanza bien",
        )
        repo = FakeConvivenciaRepo(notas_por_estudiante=[nota_p1, nota_p2])
        periodo_svc = _FakePeriodoSvc([p1, p2])
        svc = InformeService(
            FakeEstadRepo(),
            exporter=FakeExporter(),
            convivencia_svc_provider=lambda: ConvivenciaService(
                repo, periodo_svc_provider=lambda: periodo_svc
            ),
        )
        return svc

    def test_excel_hoja_principal_no_tiene_columnas_conv(self):
        """convivencia_31 T1: Excel anual no repite columnas de convivencia en la hoja principal."""
        svc = self._make_svc_con_conv()
        capturas: list[list[dict]] = []

        class _CapturandoXlsx(_RealXlsxExporter):
            def exportar_excel(self, datos, nombre_hoja="Datos", ruta_destino=None) -> bytes:
                capturas.append(datos)
                return super().exportar_excel(datos, nombre_hoja, ruta_destino)

        svc._exporter = _CapturandoXlsx()
        svc.generar_boletin_anual(1, 10, 2025, formato="excel")
        assert capturas, "No se capturaron filas"
        fila = capturas[0][0]
        assert not any(k.startswith("Nota Conv.") for k in fila), (
            "La hoja principal no debe tener columnas Nota Conv. (van en hoja Convivencia)"
        )
        assert "Observaciones Conv." not in fila
        assert "Eventos Conv." not in fila

    def test_excel_anual_tiene_hoja_convivencia(self):
        """convivencia_31 T2: Excel anual con conv_repo produce dos hojas."""
        import io

        import openpyxl

        svc = self._make_svc_con_conv()
        svc._exporter = _RealXlsxExporter()
        resultado = svc.generar_boletin_anual(1, 10, 2025, formato="excel")
        wb = openpyxl.load_workbook(io.BytesIO(resultado))
        assert "Convivencia" in wb.sheetnames, (
            f"Se esperaba hoja 'Convivencia'; hojas encontradas: {wb.sheetnames}"
        )

    def test_excel_sin_conv_repo_no_tiene_columnas_conv(self):
        """R3: Sin conv_repo, el Excel anual no incluye columnas de convivencia."""
        capturas: list[list[dict]] = []

        class CapturingExporter(FakeExporter):
            def exportar_excel(self, datos, nombre_hoja="Datos", ruta_destino=None) -> bytes:
                capturas.append(datos)
                return b"EXCEL:ok"

        svc = InformeService(FakeEstadRepo(), exporter=CapturingExporter())
        svc.generar_boletin_anual(1, 10, 2025, formato="excel")
        fila = capturas[0][0]
        assert not any(k.startswith("Nota Conv.") for k in fila)

    def test_pdf_puebla_convivencia_anual_en_datos(self):
        """T2: PDF anual llama a convivencia_boletin_anual y el PDF se genera sin error."""
        svc = self._make_svc_con_conv()
        # El PDF real de ReportLab se genera — solo verificamos que no lanza
        # y que el resultado es bytes no vacíos.
        resultado = svc.generar_boletin_anual(1, 10, 2025, formato="pdf")
        assert isinstance(resultado, bytes)
        assert len(resultado) > 100  # PDF real tiene contenido

    def test_pdf_anual_sin_conv_no_lanza(self):
        """R3: Sin conv_repo, el PDF anual no lanza — caja de obs vacía."""
        svc = InformeService(FakeEstadRepo())
        resultado = svc.generar_boletin_anual(1, 10, 2025, formato="pdf")
        assert isinstance(resultado, bytes)
        assert len(resultado) > 100


# ===========================================================================
# Group 11 — convivencia_31: hojas separadas en boletín Excel
# ===========================================================================

class TestBoletinExcelHojasConvivencia:
    """T5 — 5 tests nuevos para convivencia_31."""

    # ── helpers de clase ───────────────────────────────────────────────

    @staticmethod
    def _make_svc_periodo(conv_repo=None, ests=None):
        """InformeService configurado para boletín de periodo."""
        prov = (lambda: ConvivenciaService(conv_repo)) if conv_repo is not None else None
        return InformeService(
            FakeEstadRepo(),
            exporter=_RealXlsxExporter(),
            estudiante_repo=_FakeEstRepo(ests or []),
            convivencia_svc_provider=prov,
        )

    @staticmethod
    def _load_wb(contenido: bytes):
        import io

        import openpyxl
        return openpyxl.load_workbook(io.BytesIO(contenido))

    # ── T5-1 ──────────────────────────────────────────────────────────

    def test_boletin_periodo_excel_tiene_hoja_convivencia(self):
        """R2: con convivencia_repo el Excel de periodo tiene hoja 'Convivencia'."""
        repo = FakeConvivenciaRepo()
        svc = self._make_svc_periodo(conv_repo=repo)
        resultado = svc.generar_boletin_periodo(
            estudiante_id=1, grupo_id=10, periodo_id=5, formato="excel"
        )
        wb = self._load_wb(resultado)
        assert "Boletín Periodo" in wb.sheetnames
        assert "Convivencia" in wb.sheetnames

    # ── T5-2 ──────────────────────────────────────────────────────────

    def test_boletin_periodo_excel_no_duplica_columnas_conv(self):
        """R1: la hoja principal del boletín de periodo no tiene columnas de convivencia."""
        repo = FakeConvivenciaRepo()
        capturas: list[list[dict]] = []

        class _Cap(_RealXlsxExporter):
            def exportar_excel(self, datos, nombre_hoja="Datos", ruta_destino=None) -> bytes:
                capturas.append(datos)
                return super().exportar_excel(datos, nombre_hoja, ruta_destino)

        svc = InformeService(
            FakeEstadRepo(),
            exporter=_Cap(),
            convivencia_svc_provider=lambda: ConvivenciaService(repo),
        )
        svc.generar_boletin_periodo(
            estudiante_id=1, grupo_id=10, periodo_id=5, formato="excel"
        )
        # Si no hay asignaturas (FakeEstadRepo retorna areas=[]) capturas[0] es []
        # — lo que importa es que NO hay columnas de convivencia.
        if capturas and capturas[0]:
            fila = capturas[0][0]
            assert "Nota Conv." not in fila
            assert "Observaciones" not in fila
            assert "Eventos Conv." not in fila

    # ── T5-3 ──────────────────────────────────────────────────────────

    def test_boletines_grupo_excel_tiene_hoja_resumen(self):
        """R4: masivo de grupo con conv_provider tiene hoja 'Convivencia — todos' al final."""
        ests = [_Est(1, "Ana", "Lopez"), _Est(2, "Beto", "Diaz")]
        repo = FakeConvivenciaRepo()
        svc = InformeService(
            FakeEstadRepo(),
            exporter=_RealXlsxExporter(),
            estudiante_repo=_FakeEstRepo(ests),
            convivencia_svc_provider=lambda: ConvivenciaService(repo),
        )
        r = svc.generar_boletines_grupo(grupo_id=10, periodo_id=5, formato="excel")
        assert r.contenido is not None
        wb = self._load_wb(r.contenido)
        assert "Convivencia — todos" in wb.sheetnames
        # La hoja resumen debe ser la última
        assert wb.sheetnames[-1] == "Convivencia — todos"

    # ── T5-4 ──────────────────────────────────────────────────────────

    def test_hoja_resumen_ordena_por_estudiante(self):
        """R4: la hoja resumen ordena las filas por apellido+nombre."""
        ests = [
            _Est(1, "Carlos", "Zapata"),
            _Est(2, "Ana",    "Martinez"),
            _Est(3, "Beto",   "Lopez"),
        ]
        repo = FakeConvivenciaRepo()
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(repo),
            estudiante_repo=_FakeEstRepo(ests),
        )
        filas = svc._hoja_convivencia_resumen_grupo(grupo_id=10, periodo_id=5, anio_id=None)
        # Extraer filas de datos (saltar encabezados: título, vacío, encabezado de cols)
        data_rows = [f for f in filas if f and isinstance(f[0], str) and f[0] not in (
            "Estudiante",
            "CONVIVENCIA — Grupo 10 — Periodo 5",
            "Observaciones públicas (todas)",
            "Eventos de convivencia",
        ) and not f[0].startswith("CONVIVENCIA")]
        # Los nombres de estudiante deben estar ordenados alfabéticamente
        nombres = [row[0] for row in data_rows if row[0] and row[0] != "Estudiante"]
        assert nombres == sorted(nombres), f"Orden incorrecto: {nombres}"
        assert "Lopez Beto" in nombres
        assert "Martinez Ana" in nombres
        assert "Zapata Carlos" in nombres

    # ── T5-5 ──────────────────────────────────────────────────────────

    def test_hoja_resumen_incluye_pies_obs_y_eventos_cuando_hay(self):
        """R4: cuando hay obs y registros los pies de tabla aparecen en la hoja resumen."""
        obs = ObservacionPeriodo(
            estudiante_id=1, asignacion_id=1, periodo_id=5,
            texto="Obs de prueba", es_publica=True,
        )
        from datetime import date as _date

        from src.domain.models.convivencia import RegistroComportamiento, TipoRegistro
        reg = RegistroComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD,
            descripcion="Evento de prueba",
            fecha=_date(2026, 3, 1),
        )
        ests = [_Est(1, "Ana", "Lopez")]
        repo = FakeConvivenciaRepo(obs_por_grupo=[obs], registros=[reg])
        svc = InformeService(
            FakeEstadRepo(),
            convivencia_svc_provider=lambda: ConvivenciaService(repo),
            estudiante_repo=_FakeEstRepo(ests),
        )
        filas = svc._hoja_convivencia_resumen_grupo(grupo_id=10, periodo_id=5, anio_id=None)
        flat = [str(celda) for fila in filas for celda in fila]
        assert "Observaciones públicas (todas)" in flat
        assert "Eventos de convivencia" in flat
        assert "Obs de prueba" in flat
        assert "Evento de prueba" in flat

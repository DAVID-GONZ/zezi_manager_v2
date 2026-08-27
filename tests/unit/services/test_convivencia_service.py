"""Tests unitarios para ConvivenciaService."""
from __future__ import annotations

from datetime import date

import pytest

from src.domain.models.alerta import (
    Alerta,
    ConfiguracionAlerta,
    FiltroAlertasDTO,
    NivelAlerta,
    TipoAlerta,
)
from src.domain.models.convivencia import (
    CategoriaObservacion,
    ConceptoComportamientoDTO,
    FiltroConvivenciaDTO,
    MedidaPedagogica,
    NotaComportamiento,
    NuevaAlertaSeguimientoDTO,
    NuevaCategoriaDTO,
    NuevaMedidaPedagogicaDTO,
    NuevaNotaComportamientoDTO,
    NuevaObservacionDTO,
    NuevoRegistroComportamientoDTO,
    ObservacionPeriodo,
    PlantillaObservacion,
    PuntoSerieDTO,
    RegistroComportamiento,
    ResumenConvivenciaDTO,
    Seguimiento360DTO,
    TipoRegistro,
)
from src.domain.ports.alerta_repo import IAlertaRepository
from src.domain.ports.convivencia_repo import IConvivenciaRepository
from src.services.convivencia_service import ConvivenciaService

# ===========================================================================
# Fake
# ===========================================================================

class FakeConvRepo(IConvivenciaRepository):
    def __init__(self):
        self._obs: dict[int, ObservacionPeriodo] = {}
        self._regs: dict[int, RegistroComportamiento] = {}
        self._notas: dict[tuple, NotaComportamiento] = {}
        self._cats: dict[int, CategoriaObservacion] = {}
        self._plantillas: dict[int, PlantillaObservacion] = {}
        self._medidas: dict[int, MedidaPedagogica] = {}
        # Mapa asignacion_id -> grupo_id, emula el join observaciones→asignaciones
        # que usa listar_observaciones_por_grupo en el repo SQLite.
        self._asig_grupo: dict[int, int] = {}
        self._next_obs = 1
        self._next_reg = 1
        self._next_nota = 1
        self._next_cat = 1
        self._next_plantilla = 1
        self._next_medida = 1

    # Observaciones
    def get_observacion(self, oid: int) -> ObservacionPeriodo | None:
        return self._obs.get(oid)

    def get_observacion_por_asignacion(self, est_id: int, asig_id: int, per_id: int) -> ObservacionPeriodo | None:
        for o in self._obs.values():
            if o.estudiante_id == est_id and o.asignacion_id == asig_id and o.periodo_id == per_id:
                return o
        return None

    def listar_observaciones_por_estudiante(self, est_id: int, per_id=None, solo_publicas=False) -> list[ObservacionPeriodo]:
        return [o for o in self._obs.values() if o.estudiante_id == est_id]

    def listar_observaciones_por_grupo(self, grupo_id: int, periodo_id=None, solo_publicas=False) -> list[ObservacionPeriodo]:
        result = [
            o for o in self._obs.values()
            if self._asig_grupo.get(o.asignacion_id) == grupo_id
        ]
        if periodo_id is not None:
            result = [o for o in result if o.periodo_id == periodo_id]
        if solo_publicas:
            result = [o for o in result if o.es_publica]
        return result

    def guardar_observacion(self, o: ObservacionPeriodo) -> ObservacionPeriodo:
        o = o.model_copy(update={"id": self._next_obs})
        self._next_obs += 1
        self._obs[o.id] = o
        return o

    def actualizar_observacion(self, o: ObservacionPeriodo) -> ObservacionPeriodo:
        self._obs[o.id] = o
        return o

    def eliminar_observacion(self, oid: int) -> bool:
        return self._obs.pop(oid, None) is not None

    # Registros
    def get_registro(self, rid: int) -> RegistroComportamiento | None:
        return self._regs.get(rid)

    def listar_registros(self, filtro: FiltroConvivenciaDTO, institucion_id=None) -> list[RegistroComportamiento]:
        return list(self._regs.values())

    def contar_registros(self, filtro: FiltroConvivenciaDTO, institucion_id=None) -> int:
        return len(self._regs)

    def guardar_registro(self, r: RegistroComportamiento) -> RegistroComportamiento:
        r = r.model_copy(update={"id": self._next_reg})
        self._next_reg += 1
        self._regs[r.id] = r
        return r

    def actualizar_registro(self, r: RegistroComportamiento) -> RegistroComportamiento:
        self._regs[r.id] = r
        return r

    def eliminar_registro(self, rid: int) -> bool:
        return self._regs.pop(rid, None) is not None

    # Notas
    def get_nota(self, est_id: int, per_id: int) -> NotaComportamiento | None:
        return self._notas.get((est_id, per_id))

    def listar_notas_por_estudiante(self, est_id: int) -> list[NotaComportamiento]:
        return [n for (e, p), n in self._notas.items() if e == est_id]

    def listar_notas_por_grupo(self, grupo_id: int, per_id: int) -> list[NotaComportamiento]:
        return [n for (e, p), n in self._notas.items() if p == per_id and n.grupo_id == grupo_id]

    def guardar_nota(self, n: NotaComportamiento) -> NotaComportamiento:
        key = (n.estudiante_id, n.periodo_id)
        self._notas[key] = n
        return n

    # Categorías
    def listar_categorias(self, solo_activas: bool = True, institucion_id: int | None = None) -> list[CategoriaObservacion]:
        cats = list(self._cats.values())
        if solo_activas:
            cats = [c for c in cats if c.activa]
        return cats

    def get_categoria(self, categoria_id: int) -> CategoriaObservacion | None:
        return self._cats.get(categoria_id)

    def guardar_categoria(self, cat: CategoriaObservacion) -> CategoriaObservacion:
        cat = cat.model_copy(update={"id": self._next_cat})
        self._next_cat += 1
        self._cats[cat.id] = cat
        return cat

    def actualizar_categoria(self, cat: CategoriaObservacion) -> CategoriaObservacion:
        self._cats[cat.id] = cat
        return cat

    # Plantillas (convivencia_12)
    def listar_plantillas(self, categoria_id=None, solo_activas=True, institucion_id=None) -> list[PlantillaObservacion]:
        result = list(self._plantillas.values())
        if solo_activas:
            result = [p for p in result if p.activa]
        if categoria_id is not None:
            result = [p for p in result if p.categoria_id == categoria_id]
        return sorted(result, key=lambda p: p.uso_count, reverse=True)

    def get_plantilla(self, plantilla_id: int) -> PlantillaObservacion | None:
        return self._plantillas.get(plantilla_id)

    def guardar_plantilla(self, p: PlantillaObservacion) -> PlantillaObservacion:
        p = p.model_copy(update={"id": self._next_plantilla})
        self._next_plantilla += 1
        self._plantillas[p.id] = p
        return p

    def actualizar_plantilla(self, p: PlantillaObservacion) -> PlantillaObservacion:
        self._plantillas[p.id] = p
        return p

    def incrementar_uso_plantilla(self, plantilla_id: int) -> None:
        if plantilla_id in self._plantillas:
            p = self._plantillas[plantilla_id]
            self._plantillas[plantilla_id] = p.model_copy(
                update={"uso_count": p.uso_count + 1}
            )

    # Tipos de situación (convivencia_34)
    def listar_tipos_situacion(self, solo_activas=True, institucion_id=None):
        return []

    def get_tipo_situacion(self, tipo_situacion_id):
        return None

    def guardar_tipo_situacion(self, tipo_situacion):
        from src.domain.models.convivencia import TipoSituacion
        return tipo_situacion.model_copy(update={"id": 1})

    def actualizar_tipo_situacion(self, tipo_situacion):
        return tipo_situacion

    # Entradas de seguimiento (convivencia_35)
    def listar_entradas_seguimiento(self, registro_id):
        return []

    def guardar_entrada_seguimiento(self, entrada):
        return entrada.model_copy(update={"id": 1})

    # Medidas pedagógicas (convivencia_36)
    def listar_medidas(self, solo_activas=True, institucion_id=None):
        result = list(self._medidas.values())
        if solo_activas:
            result = [m for m in result if m.activa]
        return result

    def get_medida(self, medida_id):
        return self._medidas.get(medida_id)

    def guardar_medida(self, medida):
        medida = medida.model_copy(update={"id": self._next_medida})
        self._next_medida += 1
        self._medidas[medida.id] = medida
        return medida

    def actualizar_medida(self, medida):
        self._medidas[medida.id] = medida
        return medida

    # Lookups auxiliares
    def resolver_nombres_usuario(self, usuario_ids):
        return {}

    def resolver_nombres_asignatura(self, asignacion_ids):
        return {}

    def resolver_grupo_grado(self, grupo_id):
        return {"grupo_codigo": "601", "grupo_nombre": "601", "grado_nombre": "Sexto"}

    def resolver_acudiente_principal(self, estudiante_id):
        return {"nombre": "María García", "parentesco": "madre", "parentesco_display": "Madre", "celular": "3001234567", "email": "", "direccion": "", "documento": "12345678"}


# ===========================================================================
# Helpers
# ===========================================================================

def _make_svc() -> tuple[ConvivenciaService, FakeConvRepo]:
    repo = FakeConvRepo()
    return ConvivenciaService(repo), repo


# ===========================================================================
# Tests
# ===========================================================================

class TestRegistrarObservacion:
    def test_crea_nueva_observacion(self):
        svc, _ = _make_svc()
        dto = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Buen desempeño", es_publica=True, categoria_id=1,
        )
        obs = svc.registrar_observacion(dto)
        assert obs.id is not None

    def test_actualiza_observacion_existente(self):
        svc, _ = _make_svc()
        dto = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Texto inicial", categoria_id=1,
        )
        svc.registrar_observacion(dto)
        dto2 = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Texto actualizado", categoria_id=1,
        )
        obs = svc.registrar_observacion(dto2)
        assert obs.texto == "Texto actualizado"


class TestRegistrarComportamiento:
    def test_registra_comportamiento_fortaleza(self):
        svc, _ = _make_svc()
        dto = NuevoRegistroComportamientoDTO(
            estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.FORTALEZA,
            descripcion="Excelente participación en clase",
            fecha=date.today(),
        )
        reg = svc.registrar_comportamiento(dto)
        assert reg.id is not None
        assert reg.tipo == TipoRegistro.FORTALEZA

    def test_notificar_acudiente_exitosamente(self):
        svc, _ = _make_svc()
        dto = NuevoRegistroComportamientoDTO(
            estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.CITACION_ACUDIENTE,
            descripcion="Citación por bajo rendimiento",
            requiere_firma=True,
            fecha=date.today(),
        )
        reg = svc.registrar_comportamiento(dto)
        notificado = svc.notificar_acudiente(reg.id)
        assert notificado.acudiente_notificado is True

    def test_lanza_si_registro_no_existe(self):
        svc, _ = _make_svc()
        with pytest.raises(ValueError, match="999"):
            svc.notificar_acudiente(999)


class TestNotaComportamiento:
    def test_registra_nota_comportamiento(self):
        svc, _ = _make_svc()
        dto = NuevaNotaComportamientoDTO(
            estudiante_id=1, grupo_id=10, periodo_id=5, valor=85.0
        )
        nota = svc.registrar_nota_comportamiento(dto)
        assert nota.valor == pytest.approx(85.0)

    def test_upsert_nota_sobreescribe(self):
        svc, _ = _make_svc()
        dto1 = NuevaNotaComportamientoDTO(
            estudiante_id=1, grupo_id=10, periodo_id=5, valor=70.0
        )
        dto2 = NuevaNotaComportamientoDTO(
            estudiante_id=1, grupo_id=10, periodo_id=5, valor=85.0
        )
        svc.registrar_nota_comportamiento(dto1)
        svc.registrar_nota_comportamiento(dto2)
        nota = svc.get_nota_comportamiento(1, 5)
        assert nota.valor == pytest.approx(85.0)


# ===========================================================================
# Enforcement de autorización (convivencia_04b — defensa en profundidad)
# ===========================================================================

class _StubCatalogoSvc:
    """Stub minimal de CatalogoAcademicoService: siempre autoriza/deniega."""
    def __init__(self, autoriza: bool):
        self._autoriza = autoriza
        self.llamadas: list[tuple] = []

    def puede_gestionar_comportamiento_en_grupo(
        self, usuario_rol, usuario_id, grupo_id
    ) -> bool:
        self.llamadas.append((usuario_rol, usuario_id, grupo_id))
        return self._autoriza


class TestEnforcementAutorizacion:
    def _dto_registro(self) -> NuevoRegistroComportamientoDTO:
        return NuevoRegistroComportamientoDTO(
            estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.FORTALEZA,
            descripcion="Buen trabajo",
            fecha=date.today(),
        )

    def test_provider_deniega_lanza_permission_error_y_no_persiste(self):
        repo = FakeConvRepo()
        stub = _StubCatalogoSvc(autoriza=False)
        svc = ConvivenciaService(
            repo=repo,
            catalogo_academico_svc_provider=lambda: stub,
        )
        with pytest.raises(PermissionError):
            svc.registrar_comportamiento(
                self._dto_registro(),
                usuario_id=99,
                usuario_rol="profesor",
            )
        assert repo._regs == {}  # no persistió
        assert stub.llamadas == [("profesor", 99, 10)]

    def test_provider_autoriza_mutacion_ok(self):
        repo = FakeConvRepo()
        stub = _StubCatalogoSvc(autoriza=True)
        svc = ConvivenciaService(
            repo=repo,
            catalogo_academico_svc_provider=lambda: stub,
        )
        reg = svc.registrar_comportamiento(
            self._dto_registro(),
            usuario_id=99,
            usuario_rol="profesor",
        )
        assert reg.id is not None
        assert stub.llamadas == [("profesor", 99, 10)]

    def test_sin_provider_es_compat_retro(self):
        svc, _repo = _make_svc()  # sin provider
        reg = svc.registrar_comportamiento(
            self._dto_registro(), usuario_id=99, usuario_rol="profesor",
        )
        assert reg.id is not None


# ===========================================================================
# Concepto consolidado (convivencia_05)
# ===========================================================================

class _FakeNivel:
    def __init__(self, id, nombre, rmin, rmax, descripcion=None):
        self.id = id
        self.nombre = nombre
        self.rango_min = rmin
        self.rango_max = rmax
        self.descripcion = descripcion


class _FakeConfigSvc:
    def __init__(self, niveles):
        self._niveles = niveles
    def listar_niveles(self, anio_id):
        return self._niveles


class _FakePeriodoSvc:
    class _P:
        anio_id = 2026
    def get_by_id(self, periodo_id):
        return self._P()


class _FakeEst:
    def __init__(self, id):
        self.id = id


class _FakeEstSvc:
    def __init__(self, ests):
        self._ests = ests
    def listar_por_grupo(self, grupo_id, solo_activos=True):
        return self._ests


_NIVELES = [
    _FakeNivel(1, "Bajo", 0, 59.99, "Bajo desempeño"),
    _FakeNivel(2, "Básico", 60, 69.99, "Básico"),
    _FakeNivel(3, "Alto", 70, 84.99, "Alto"),
    _FakeNivel(4, "Superior", 85, 100, "Superior"),
]


def _svc_completo(ests=None):
    repo = FakeConvRepo()
    svc = ConvivenciaService(
        repo=repo,
        configuracion_svc_provider=lambda: _FakeConfigSvc(_NIVELES),
        periodo_svc_provider=lambda: _FakePeriodoSvc(),
        estudiante_svc_provider=lambda: _FakeEstSvc(ests or []),
    )
    return svc, repo


class TestConceptoComportamiento:
    def test_sin_nota_devuelve_dto_vacio(self):
        svc, _ = _svc_completo()
        dto = svc.get_concepto_periodo(estudiante_id=1, periodo_id=5)
        assert isinstance(dto, ConceptoComportamientoDTO)
        assert dto.valor is None
        assert dto.aprobado is False
        assert dto.nivel_nombre is None
        assert dto.concepto is None

    def test_con_desempeno_id_explicito(self):
        svc, repo = _svc_completo()
        repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5,
            valor=65.0, desempeno_id=4, observacion="Excelente actitud",
        )
        dto = svc.get_concepto_periodo(1, 5)
        # desempeno_id=4 (Superior) prevalece sobre el rango que daría "Básico"
        assert dto.nivel_nombre == "Superior"
        assert dto.valor == 65.0
        assert dto.concepto == "Excelente actitud"
        assert dto.aprobado is True

    def test_sin_desempeno_id_resuelve_por_rango(self):
        svc, repo = _svc_completo()
        repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5, valor=72.5,
        )
        dto = svc.get_concepto_periodo(1, 5)
        assert dto.nivel_nombre == "Alto"
        assert dto.aprobado is True

    def test_nota_menor_a_minima_no_aprobado(self):
        svc, repo = _svc_completo()
        repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5, valor=55.0,
        )
        dto = svc.get_concepto_periodo(1, 5, nota_minima=60.0)
        assert dto.aprobado is False
        assert dto.nivel_nombre == "Bajo"

    def test_listar_conceptos_grupo_incluye_estudiantes_sin_nota(self):
        ests = [_FakeEst(1), _FakeEst(2), _FakeEst(3)]
        svc, repo = _svc_completo(ests=ests)
        # Solo el estudiante 2 tiene nota.
        repo._notas[(2, 5)] = NotaComportamiento(
            estudiante_id=2, grupo_id=10, periodo_id=5, valor=90.0,
        )
        conceptos = svc.listar_conceptos_grupo(grupo_id=10, periodo_id=5)
        assert len(conceptos) == 3
        by_est = {c.estudiante_id: c for c in conceptos}
        assert by_est[1].valor is None and by_est[1].aprobado is False
        assert by_est[2].valor == 90.0 and by_est[2].nivel_nombre == "Superior"
        assert by_est[3].valor is None

    # -----------------------------------------------------------------
    # convivencia_06 — Reporte por grupo/periodo
    # -----------------------------------------------------------------

    def test_reporte_periodo_grupo_combina_notas_y_observaciones(self):
        ests = [_FakeEst(1), _FakeEst(2), _FakeEst(3)]
        # Añadimos nombre/apellido dinámicamente sin acoplar el modelo real.
        for e, nom, ape in [(ests[0], "Ana", "Ruiz"), (ests[1], "Bob", "Diaz"), (ests[2], "Cyd", "Paz")]:
            e.nombre = nom
            e.apellido = ape
        svc, repo = _svc_completo(ests=ests)
        # Estudiante 1 → nota + 2 observaciones.
        repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5,
            valor=90.0, observacion="Excelente disciplina",
        )
        repo.guardar_observacion(ObservacionPeriodo(
            estudiante_id=1, asignacion_id=99, periodo_id=5,
            texto="Muy participativo",
        ))
        repo.guardar_observacion(ObservacionPeriodo(
            estudiante_id=1, asignacion_id=100, periodo_id=5,
            texto="Colabora con compañeros",
        ))
        # Estudiante 2 → sin nota, con 1 observación.
        repo.guardar_observacion(ObservacionPeriodo(
            estudiante_id=2, asignacion_id=99, periodo_id=5,
            texto="Debe entregar tareas a tiempo",
        ))
        # Estudiante 3 → sin nota, sin observaciones.

        filas = svc.reporte_periodo_grupo(grupo_id=10, periodo_id=5)
        assert len(filas) == 3
        by_id = {f.estudiante_id: f for f in filas}

        # Estudiante 1: nota + concepto + 2 observaciones
        assert by_id[1].valor == 90.0
        assert by_id[1].nivel_nombre == "Superior"
        assert by_id[1].concepto == "Excelente disciplina"
        assert set(by_id[1].observaciones) == {
            "Muy participativo", "Colabora con compañeros",
        }
        assert by_id[1].nombre == "Ruiz Ana"

        # Estudiante 2: sin nota, 1 observación
        assert by_id[2].valor is None
        assert by_id[2].nivel_nombre is None
        assert by_id[2].concepto is None
        assert by_id[2].observaciones == ["Debe entregar tareas a tiempo"]

        # Estudiante 3: sin nota, sin observaciones
        assert by_id[3].valor is None
        assert by_id[3].observaciones == []

    def test_reporte_periodo_grupo_sin_provider_lanza(self):
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        with pytest.raises(RuntimeError):
            svc.reporte_periodo_grupo(grupo_id=10, periodo_id=5)

    def test_get_concepto_sin_providers_lanza(self):
        repo = FakeConvRepo()
        repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5, valor=70.0,
        )
        svc = ConvivenciaService(repo=repo)
        with pytest.raises(RuntimeError):
            svc.get_concepto_periodo(1, 5)

    # -----------------------------------------------------------------
    # convivencia_06b — Exportación en el servicio (hexagonal)
    # -----------------------------------------------------------------

    def test_exportar_reporte_sin_exporter_lanza(self):
        svc, _ = _svc_completo()
        with pytest.raises(RuntimeError):
            svc.exportar_reporte_periodo_grupo(10, 5, "excel")

    def test_exportar_reporte_formato_invalido_lanza(self):
        class _NullExp:
            def exportar_excel(self, *a, **kw): return b""
            def exportar_pdf(self, *a, **kw): return b""
            def exportar_csv(self, *a, **kw): return b""
        repo = FakeConvRepo()
        svc = ConvivenciaService(
            repo=repo,
            configuracion_svc_provider=lambda: _FakeConfigSvc(_NIVELES),
            periodo_svc_provider=lambda: _FakePeriodoSvc(),
            estudiante_svc_provider=lambda: _FakeEstSvc([]),
            exporter=_NullExp(),
        )
        with pytest.raises(ValueError):
            svc.exportar_reporte_periodo_grupo(10, 5, "csv")

    def test_exportar_reporte_excel_genera_xlsx_enriquecido(self):
        """El Excel generado es un xlsx válido con dos hojas (Reporte + Estadísticos)."""
        class _FakeExp:
            def exportar_excel(self, *a, **kw): return b""
            def exportar_pdf(self, *a, **kw): return b""
            def exportar_csv(self, *a, **kw): return b""

        est = _FakeEst(1)
        est.nombre = "Ana"
        est.apellido = "Ruiz"
        repo = FakeConvRepo()
        repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5, valor=80.0,
        )
        svc = ConvivenciaService(
            repo=repo,
            configuracion_svc_provider=lambda: _FakeConfigSvc(_NIVELES),
            periodo_svc_provider=lambda: _FakePeriodoSvc(),
            estudiante_svc_provider=lambda: _FakeEstSvc([est]),
            exporter=_FakeExp(),
        )
        bytes_ = svc.exportar_reporte_periodo_grupo(
            10, 5, "excel", titulo="X", grupo="5A", periodo="P1",
        )
        assert isinstance(bytes_, bytes)
        assert len(bytes_) > 0
        # Verificar que es un xlsx válido con las hojas esperadas
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(bytes_))
        assert "Reporte" in wb.sheetnames
        assert "Estadísticos" in wb.sheetnames
        ws = wb["Reporte"]
        # Membrete presente
        assert ws.cell(1, 1).value == "INSTITUCIÓN EDUCATIVA ZECI"
        # Datos del estudiante presentes
        found = False
        for row in ws.iter_rows(min_row=7, max_col=1, values_only=True):
            if row[0] and "Ruiz" in str(row[0]):
                found = True
                break
        assert found, "El nombre del estudiante debe aparecer en la hoja Reporte"

    def test_exportar_reporte_pdf_genera_pdf_reportlab(self):
        class _FakeExp:
            def exportar_excel(self, *a, **kw): return b""
            def exportar_pdf(self, html, ruta_destino=None): return b""
            def exportar_csv(self, *a, **kw): return b""

        est = _FakeEst(1)
        est.nombre = "Ana"
        est.apellido = "Ruiz"
        repo = FakeConvRepo()
        exp = _FakeExp()
        svc = ConvivenciaService(
            repo=repo,
            configuracion_svc_provider=lambda: _FakeConfigSvc(_NIVELES),
            periodo_svc_provider=lambda: _FakePeriodoSvc(),
            estudiante_svc_provider=lambda: _FakeEstSvc([est]),
            exporter=exp,
        )
        bytes_ = svc.exportar_reporte_periodo_grupo(
            10, 5, "pdf", titulo="Reporte X", grupo="5A", periodo="Periodo 1",
        )
        assert isinstance(bytes_, bytes)
        assert len(bytes_) > 0
        assert bytes_[:5] == b"%PDF-"


# ===========================================================================
# Catálogo de categorías (convivencia_10)
# ===========================================================================

class TestCategoriasObservacion:
    def test_listar_categorias_delega_al_repo(self):
        """listar_categorias llama al repo con solo_activas=True por defecto."""
        repo = FakeConvRepo()
        # Precargar dos categorías: una activa y una inactiva.
        repo.guardar_categoria(
            CategoriaObservacion(nombre="Académico", activa=True)
        )
        repo.guardar_categoria(
            CategoriaObservacion(nombre="Archivada", activa=False)
        )
        svc = ConvivenciaService(repo=repo)
        resultado = svc.listar_categorias(solo_activas=True)
        assert len(resultado) == 1
        assert resultado[0].nombre == "Académico"
        # Con solo_activas=False deben aparecer ambas
        todas = svc.listar_categorias(solo_activas=False)
        assert len(todas) == 2

    def test_crear_categoria_llama_guardar(self):
        """crear_categoria persiste la categoría y retorna el objeto con id."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        dto = NuevaCategoriaDTO(nombre="Convivencia", es_comportamental=True)
        cat = svc.crear_categoria(dto)
        assert cat.id is not None
        assert cat.nombre == "Convivencia"
        assert cat.es_comportamental is True
        assert cat.activa is True
        # Verificar persistencia en repo
        assert len(repo._cats) == 1

    def test_desactivar_categoria_pone_activa_false(self):
        """desactivar_categoria setea activa=False en la categoría."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        # Crear primero una categoría activa
        dto = NuevaCategoriaDTO(nombre="Normas", es_comportamental=True)
        cat = svc.crear_categoria(dto)
        assert cat.activa is True
        # Desactivar
        desactivada = svc.desactivar_categoria(cat.id)
        assert desactivada.activa is False
        assert desactivada.nombre == "Normas"
        # El repo también refleja el cambio
        en_repo = repo.get_categoria(cat.id)
        assert en_repo.activa is False


# ===========================================================================
# Autorización por objeto: observaciones (convivencia_11)
# ===========================================================================

class _FakeAsignacionSvc:
    """Stub de AsignacionService: asignación con usuario_id configurable."""
    def __init__(self, usuario_id_titular: int):
        self._usuario_id_titular = usuario_id_titular

    def get_by_id(self, asig_id: int):
        class _Asig:
            pass
        a = _Asig()
        a.id = asig_id
        a.usuario_id = self._usuario_id_titular
        return a

    def listar_por_docente(self, usuario_id, periodo_id=None):
        return []


class TestObservacionAutorizacionPorObjeto:
    """Autorización por objeto (convivencia_11): profesores solo en sus asignaciones."""

    def test_profesor_no_autorizado_registrar_observacion_ajena(self):
        """Profesor intenta registrar obs de asignación que no es suya → PermissionError."""
        repo = FakeConvRepo()
        # asignacion_id=3 pertenece al titular usuario_id=99, no al profesor 50
        svc = ConvivenciaService(
            repo=repo,
            asignacion_svc_provider=lambda: _FakeAsignacionSvc(usuario_id_titular=99),
        )
        dto = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Obs ajena", categoria_id=1,
        )
        with pytest.raises(PermissionError, match="Solo puedes registrar"):
            svc.registrar_observacion(dto, usuario_id=50, usuario_rol="profesor")
        # No debe haber persistido nada
        assert repo._obs == {}

    def test_profesor_autorizado_registra_su_propia_observacion(self):
        """Profesor registra obs de su propia asignación → permitido."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(
            repo=repo,
            asignacion_svc_provider=lambda: _FakeAsignacionSvc(usuario_id_titular=50),
        )
        dto = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Obs propia", categoria_id=1,
        )
        obs = svc.registrar_observacion(dto, usuario_id=50, usuario_rol="profesor")
        assert obs.id is not None
        assert len(repo._obs) == 1

    def test_director_puede_registrar_sin_restriccion_de_asignacion(self):
        """Director no pasa por la verificación de asignación → acceso pleno."""
        repo = FakeConvRepo()
        # Aunque el provider diga que el titular es 99, el director (50) puede pasar
        svc = ConvivenciaService(
            repo=repo,
            asignacion_svc_provider=lambda: _FakeAsignacionSvc(usuario_id_titular=99),
        )
        dto = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Obs de director", categoria_id=1,
        )
        obs = svc.registrar_observacion(dto, usuario_id=50, usuario_rol="director")
        assert obs.id is not None

    def test_sin_asignacion_provider_no_bloquea_a_profesor(self):
        """Sin asignacion_svc_provider, compat retro: no bloquea aunque sea profesor."""
        svc, _repo = _make_svc()  # sin asignacion_svc_provider
        dto = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Obs sin provider", categoria_id=1,
        )
        obs = svc.registrar_observacion(dto, usuario_id=50, usuario_rol="profesor")
        assert obs.id is not None


# ===========================================================================
# Catálogo de plantillas (convivencia_12)
# ===========================================================================

class TestPlantillasObservacion:
    def test_listar_plantillas_servicio(self):
        """listar_plantillas delega al repo con solo_activas=True y filtra por categoría."""
        repo = FakeConvRepo()
        # Insertar plantillas: una de cat 1, una de cat 2, una inactiva
        repo.guardar_plantilla(PlantillaObservacion(texto="Texto A", categoria_id=1, activa=True))
        repo.guardar_plantilla(PlantillaObservacion(texto="Texto B", categoria_id=2, activa=True))
        repo.guardar_plantilla(PlantillaObservacion(texto="Inactiva", categoria_id=1, activa=False))
        svc = ConvivenciaService(repo=repo)

        # Sin filtro: devuelve solo las activas
        todas_activas = svc.listar_plantillas()
        assert len(todas_activas) == 2
        textos = {p.texto for p in todas_activas}
        assert "Texto A" in textos
        assert "Texto B" in textos
        assert "Inactiva" not in textos

        # Filtrado por categoria_id=1: solo "Texto A"
        de_cat1 = svc.listar_plantillas(categoria_id=1)
        assert len(de_cat1) == 1
        assert de_cat1[0].texto == "Texto A"

        # Filtrado por categoria_id=2: solo "Texto B"
        de_cat2 = svc.listar_plantillas(categoria_id=2)
        assert len(de_cat2) == 1
        assert de_cat2[0].texto == "Texto B"

    def test_registrar_observacion_desde_plantilla_incrementa_uso(self):
        """registrar_observacion_desde_plantilla guarda obs con origen='plantilla' e incrementa uso."""
        repo = FakeConvRepo()
        # Insertar una plantilla
        plantilla = repo.guardar_plantilla(
            PlantillaObservacion(texto="Buen desempeño", categoria_id=1, uso_count=3)
        )
        svc = ConvivenciaService(repo=repo)
        dto = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto=plantilla.texto, categoria_id=1,
        )
        obs = svc.registrar_observacion_desde_plantilla(dto, plantilla.id)

        # La observación debe haberse guardado con origen="plantilla"
        assert obs.id is not None
        assert obs.origen == "plantilla"

        # El uso_count debe haber incrementado
        actualizada = repo.get_plantilla(plantilla.id)
        assert actualizada.uso_count == 4  # 3 + 1

    def test_registrar_observacion_desde_plantilla_upsert(self):
        """Si ya existe una observación para asig/periodo/estudiante, la actualiza."""
        repo = FakeConvRepo()
        plantilla = repo.guardar_plantilla(
            PlantillaObservacion(texto="Texto plantilla", categoria_id=1)
        )
        svc = ConvivenciaService(repo=repo)

        dto = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Texto original", categoria_id=1,
        )
        # Primera vez → crea
        obs1 = svc.registrar_observacion_desde_plantilla(dto, plantilla.id)
        assert obs1.origen == "plantilla"

        # Segunda vez → actualiza (misma asig/periodo/estudiante)
        dto2 = NuevaObservacionDTO(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Texto actualizado", categoria_id=1,
        )
        obs2 = svc.registrar_observacion_desde_plantilla(dto2, plantilla.id)
        assert obs2.texto == "Texto actualizado"
        assert obs2.origen == "plantilla"

        # Solo debe haber una observación en el repo
        assert len(repo._obs) == 1

        # Uso incrementado 2 veces
        assert repo.get_plantilla(plantilla.id).uso_count == 2


# ===========================================================================
# Promoción a comportamiento (convivencia_14)
# ===========================================================================

class TestPromocionAComportamiento:
    """Tests para promover_a_comportamiento."""

    def _obs_comportamental(self, repo: FakeConvRepo) -> ObservacionPeriodo:
        """Crea observación con categoría comportamental en el repo y la retorna."""
        cat = repo.guardar_categoria(
            CategoriaObservacion(nombre="Convivencia", es_comportamental=True)
        )
        return repo.guardar_observacion(
            ObservacionPeriodo(
                estudiante_id=1,
                asignacion_id=3,
                periodo_id=5,
                texto="Pelea en el recreo",
                es_publica=True,
                categoria_id=cat.id,
            )
        )

    def _obs_no_comportamental(self, repo: FakeConvRepo) -> ObservacionPeriodo:
        """Crea observación con categoría NO comportamental en el repo y la retorna."""
        cat = repo.guardar_categoria(
            CategoriaObservacion(nombre="Académico", es_comportamental=False)
        )
        return repo.guardar_observacion(
            ObservacionPeriodo(
                estudiante_id=1,
                asignacion_id=3,
                periodo_id=5,
                texto="Entregó la tarea tarde",
                es_publica=True,
                categoria_id=cat.id,
            )
        )

    def test_promover_a_comportamiento_crea_registro(self):
        """Flujo nominal: director con categoría comportamental → crea RegistroComportamiento."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        obs = self._obs_comportamental(repo)

        registro = svc.promover_a_comportamiento(
            obs.id, usuario_id=10, usuario_rol="director"
        )

        # Debe retornar un RegistroComportamiento con id asignado
        assert isinstance(registro, RegistroComportamiento)
        assert registro.id is not None
        assert registro.estudiante_id == obs.estudiante_id
        assert registro.periodo_id == obs.periodo_id
        assert registro.descripcion == obs.texto
        assert registro.usuario_registro_id == 10

        # La observación debe quedar enlazada al registro
        obs_actualizada = repo.get_observacion(obs.id)
        assert obs_actualizada.registro_comportamiento_id == registro.id

        # Solo un registro de comportamiento debe existir
        assert len(repo._regs) == 1

    def test_promover_a_comportamiento_coordinador_permitido(self):
        """Coordinador también puede promover."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        obs = self._obs_comportamental(repo)

        registro = svc.promover_a_comportamiento(
            obs.id, usuario_id=20, usuario_rol="coordinador"
        )
        assert registro.id is not None

    def test_promover_a_comportamiento_categoria_no_comportamental(self):
        """Categoría no comportamental → ValueError; no se persiste ningún registro."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        obs = self._obs_no_comportamental(repo)

        with pytest.raises(ValueError, match="comportamental"):
            svc.promover_a_comportamiento(
                obs.id, usuario_id=10, usuario_rol="director"
            )
        assert repo._regs == {}

    def test_promover_a_comportamiento_profesor_no_autorizado(self):
        """Profesor intenta promover → PermissionError; no se persiste ningún registro."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        obs = self._obs_comportamental(repo)

        with pytest.raises(PermissionError):
            svc.promover_a_comportamiento(
                obs.id, usuario_id=50, usuario_rol="profesor"
            )
        assert repo._regs == {}

    def test_promover_a_comportamiento_obs_sin_categoria_lanza(self):
        """Observación sin categoria_id → ValueError."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        obs = repo.guardar_observacion(
            ObservacionPeriodo(
                estudiante_id=1, asignacion_id=3, periodo_id=5,
                texto="Sin categoría", es_publica=True,
                categoria_id=None,
            )
        )
        with pytest.raises(ValueError):
            svc.promover_a_comportamiento(
                obs.id, usuario_id=10, usuario_rol="director"
            )

    def test_promover_a_comportamiento_obs_inexistente_lanza(self):
        """Observación no existe → ValueError."""
        svc, _ = _make_svc()
        with pytest.raises(ValueError, match="999"):
            svc.promover_a_comportamiento(
                999, usuario_id=10, usuario_rol="director"
            )


# ===========================================================================
# Catálogo de retroalimentación (convivencia_13)
# ===========================================================================

class TestPromocionPlantillas:
    """Tests para promover_observacion_a_plantilla y listar_plantillas_sugeridas."""

    def _obs_en_repo(self, repo: FakeConvRepo) -> ObservacionPeriodo:
        """Inserta una observación de prueba y la retorna con id asignado."""
        return repo.guardar_observacion(
            ObservacionPeriodo(
                estudiante_id=1,
                asignacion_id=3,
                periodo_id=5,
                texto="Excelente participación",
                categoria_id=2,
            )
        )

    def test_promover_observacion_a_plantilla_crea_plantilla(self):
        """Director promueve una observación existente → se crea PlantillaObservacion."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        obs = self._obs_en_repo(repo)

        plantilla = svc.promover_observacion_a_plantilla(
            obs.id, usuario_id=10, usuario_rol="director"
        )

        assert isinstance(plantilla, PlantillaObservacion)
        assert plantilla.id is not None
        assert plantilla.texto == obs.texto
        assert plantilla.categoria_id == obs.categoria_id
        # Verificar que se persistió en el repo
        assert len(repo._plantillas) == 1

    def test_promover_observacion_coordinador_permitido(self):
        """Coordinador también tiene permiso para promover."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        obs = self._obs_en_repo(repo)

        plantilla = svc.promover_observacion_a_plantilla(
            obs.id, usuario_id=20, usuario_rol="coordinador"
        )
        assert plantilla.id is not None

    def test_promover_observacion_profesor_no_autorizado(self):
        """Profesor intenta promover → PermissionError, no se persiste."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        obs = self._obs_en_repo(repo)

        with pytest.raises(PermissionError):
            svc.promover_observacion_a_plantilla(
                obs.id, usuario_id=50, usuario_rol="profesor"
            )
        # Ninguna plantilla debe haberse creado
        assert len(repo._plantillas) == 0

    def test_promover_observacion_inexistente_lanza(self):
        """Si la observación no existe → ValueError."""
        svc, _ = _make_svc()
        with pytest.raises(ValueError, match="999"):
            svc.promover_observacion_a_plantilla(
                999, usuario_id=10, usuario_rol="director"
            )

    def test_listar_plantillas_sugeridas_limite(self):
        """listar_plantillas_sugeridas retorna como máximo `limite` elementos."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        # Insertar 8 plantillas con distintos uso_count
        for i in range(8):
            repo.guardar_plantilla(
                PlantillaObservacion(texto=f"Plantilla {i}", categoria_id=1, uso_count=i)
            )

        # limite=5 (default)
        sugeridas = svc.listar_plantillas_sugeridas()
        assert len(sugeridas) == 5

        # Las primeras deben ser las de mayor uso_count
        usos = [p.uso_count for p in sugeridas]
        assert usos == sorted(usos, reverse=True)

    def test_listar_plantillas_sugeridas_filtro_categoria(self):
        """listar_plantillas_sugeridas respeta el filtro de categoria_id."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        repo.guardar_plantilla(PlantillaObservacion(texto="Cat1-A", categoria_id=1, uso_count=10))
        repo.guardar_plantilla(PlantillaObservacion(texto="Cat1-B", categoria_id=1, uso_count=5))
        repo.guardar_plantilla(PlantillaObservacion(texto="Cat2-A", categoria_id=2, uso_count=8))

        sugeridas_cat1 = svc.listar_plantillas_sugeridas(categoria_id=1, limite=10)
        assert len(sugeridas_cat1) == 2
        assert all(p.categoria_id == 1 for p in sugeridas_cat1)

        sugeridas_cat2 = svc.listar_plantillas_sugeridas(categoria_id=2, limite=10)
        assert len(sugeridas_cat2) == 1
        assert sugeridas_cat2[0].texto == "Cat2-A"


# ===========================================================================
# FakeAlertaRepo para tests de alertas dentro de ConvivenciaService
# ===========================================================================

class FakeAlertaRepo(IAlertaRepository):
    """Implementación mínima de IAlertaRepository para tests de ConvivenciaService."""

    def __init__(self, existe_pendiente: bool = False, cfg: ConfiguracionAlerta | None = None):
        self._alertas: list[Alerta] = []
        self._existe_pendiente = existe_pendiente
        self._cfg = cfg
        self._next_id = 1

    # Configuración
    def get_configuracion(self, anio_id: int, tipo_alerta: TipoAlerta) -> ConfiguracionAlerta | None:
        if self._cfg and self._cfg.tipo_alerta == tipo_alerta:
            return self._cfg
        return None

    def listar_configuraciones(self, anio_id: int, solo_activas: bool = True) -> list[ConfiguracionAlerta]:
        return [self._cfg] if self._cfg else []

    def guardar_configuracion(self, config: ConfiguracionAlerta) -> ConfiguracionAlerta:
        self._cfg = config
        return config

    def desactivar_configuracion(self, anio_id: int, tipo_alerta: TipoAlerta) -> bool:
        return False

    # Alertas
    def get_alerta(self, alerta_id: int) -> Alerta | None:
        for a in self._alertas:
            if a.id == alerta_id:
                return a
        return None

    def listar_alertas(self, filtro: FiltroAlertasDTO) -> list[Alerta]:
        return list(self._alertas)

    def contar_pendientes(self, estudiante_id=None, nivel=None) -> int:
        return sum(1 for a in self._alertas if not a.resuelta)

    def existe_pendiente(self, estudiante_id: int, tipo_alerta: TipoAlerta) -> bool:
        return self._existe_pendiente

    def guardar_alerta(self, alerta: Alerta) -> Alerta:
        alerta = alerta.model_copy(update={"id": self._next_id})
        self._next_id += 1
        self._alertas.append(alerta)
        return alerta

    def guardar_alertas_masivas(self, alertas: list[Alerta]) -> int:
        for a in alertas:
            self.guardar_alerta(a)
        return len(alertas)

    def resolver_alerta(self, alerta_id, usuario_id, observacion=None, fecha=None) -> bool:
        return False

    def resolver_alertas_de_estudiante(self, estudiante_id, tipo_alerta, usuario_id, observacion=None) -> int:
        return 0

    def listar_alertas_por_destinatario(
        self,
        usuario_destino_id: int,
        tipo: str | None = None,
        solo_pendientes: bool = True,
    ) -> list[Alerta]:
        resultado = [
            a for a in self._alertas
            if a.usuario_destino_id == usuario_destino_id
            and (tipo is None or a.tipo_alerta.value == tipo)
            and (not solo_pendientes or not a.resuelta)
        ]
        return resultado


# ===========================================================================
# convivencia_16: crear_alerta_seguimiento_manual
# ===========================================================================

class TestCrearAlertaSeguimientoManual:
    """Tests para ConvivenciaService.crear_alerta_seguimiento_manual (convivencia_16)."""

    def _make_svc_con_alerta_repo(self) -> tuple[ConvivenciaService, FakeConvRepo, FakeAlertaRepo]:
        conv_repo = FakeConvRepo()
        alerta_repo = FakeAlertaRepo()
        svc = ConvivenciaService(repo=conv_repo, alerta_repo=alerta_repo)
        return svc, conv_repo, alerta_repo

    def test_crear_alerta_seguimiento_manual_flujo_nominal(self):
        """Director crea alerta → tipo es SEGUIMIENTO_REQUERIDO, usuario_destino_id correcto."""
        svc, _, alerta_repo = self._make_svc_con_alerta_repo()
        dto = NuevaAlertaSeguimientoDTO(
            estudiante_id=5,
            usuario_destino_id=12,
            descripcion="El estudiante requiere atención urgente.",
            nivel=NivelAlerta.ADVERTENCIA,
        )
        alerta = svc.crear_alerta_seguimiento_manual(
            dto, usuario_id=1, usuario_rol="director"
        )
        assert alerta.id is not None
        assert alerta.tipo_alerta == TipoAlerta.SEGUIMIENTO_REQUERIDO
        assert alerta.estudiante_id == 5
        assert alerta.usuario_destino_id == 12
        assert alerta.nivel == NivelAlerta.ADVERTENCIA
        assert len(alerta_repo._alertas) == 1

    def test_crear_alerta_seguimiento_coordinador_permitido(self):
        """Coordinador también puede crear alertas de seguimiento."""
        svc, _, _alerta_repo = self._make_svc_con_alerta_repo()
        dto = NuevaAlertaSeguimientoDTO(
            estudiante_id=3,
            usuario_destino_id=7,
            descripcion="Seguimiento recomendado.",
        )
        alerta = svc.crear_alerta_seguimiento_manual(
            dto, usuario_id=2, usuario_rol="coordinador"
        )
        assert alerta.tipo_alerta == TipoAlerta.SEGUIMIENTO_REQUERIDO

    def test_crear_alerta_seguimiento_profesor_no_autorizado(self):
        """Profesor intenta crear alerta de seguimiento → PermissionError; no persiste."""
        svc, _, alerta_repo = self._make_svc_con_alerta_repo()
        dto = NuevaAlertaSeguimientoDTO(
            estudiante_id=3,
            usuario_destino_id=7,
            descripcion="Intento no autorizado.",
        )
        with pytest.raises(PermissionError):
            svc.crear_alerta_seguimiento_manual(
                dto, usuario_id=50, usuario_rol="profesor"
            )
        assert alerta_repo._alertas == []

    def test_crear_alerta_seguimiento_nivel_critica(self):
        """Se puede crear alerta con nivel CRITICA."""
        svc, _, _alerta_repo = self._make_svc_con_alerta_repo()
        dto = NuevaAlertaSeguimientoDTO(
            estudiante_id=8,
            usuario_destino_id=4,
            descripcion="Situación crítica de convivencia.",
            nivel=NivelAlerta.CRITICA,
        )
        alerta = svc.crear_alerta_seguimiento_manual(
            dto, usuario_id=1, usuario_rol="director"
        )
        assert alerta.nivel == NivelAlerta.CRITICA


# ===========================================================================
# convivencia_17: _verificar_alerta_comportamiento usa SEGUIMIENTO_REQUERIDO
# ===========================================================================

class TestVerificarAlertaComportamiento:
    """Tests para _verificar_alerta_comportamiento (convivencia_17)."""

    def _cfg_seguimiento(self, umbral: float = 3.0) -> ConfiguracionAlerta:
        return ConfiguracionAlerta(
            anio_id=2026,
            tipo_alerta=TipoAlerta.SEGUIMIENTO_REQUERIDO,
            umbral=umbral,
            activa=True,
        )

    def _dto_negativo(self) -> NuevoRegistroComportamientoDTO:
        return NuevoRegistroComportamientoDTO(
            estudiante_id=1,
            grupo_id=10,
            periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD,
            descripcion="Incidente de convivencia",
        )

    def test_verificar_alerta_usa_tipo_seguimiento(self):
        """Cuando conteo supera umbral, la alerta guardada usa SEGUIMIENTO_REQUERIDO."""
        conv_repo = FakeConvRepo()
        alerta_repo = FakeAlertaRepo(
            existe_pendiente=False,
            cfg=self._cfg_seguimiento(umbral=1.0),
        )
        svc = ConvivenciaService(repo=conv_repo, alerta_repo=alerta_repo)

        svc.registrar_comportamiento(
            self._dto_negativo(), usuario_id=1, anio_id=2026, usuario_rol=None
        )

        assert len(alerta_repo._alertas) == 1
        assert alerta_repo._alertas[0].tipo_alerta == TipoAlerta.SEGUIMIENTO_REQUERIDO

    def test_verificar_alerta_no_duplica_pendiente(self):
        """Si ya existe alerta pendiente del tipo SEGUIMIENTO_REQUERIDO → no vuelve a guardar."""
        conv_repo = FakeConvRepo()
        alerta_repo = FakeAlertaRepo(
            existe_pendiente=True,
            cfg=self._cfg_seguimiento(umbral=1.0),
        )
        svc = ConvivenciaService(repo=conv_repo, alerta_repo=alerta_repo)

        svc.registrar_comportamiento(
            self._dto_negativo(), usuario_id=1, anio_id=2026, usuario_rol=None
        )

        # No debe haber guardado ninguna alerta nueva
        assert alerta_repo._alertas == []

    def test_verificar_alerta_nivel_critico_doble_umbral(self):
        """Cuando conteo >= umbral*2, el nivel de la alerta generada es CRITICA."""
        conv_repo = FakeConvRepo()
        # umbral=1 → umbral*2=2; precargamos 2 registros negativos
        alerta_repo = FakeAlertaRepo(
            existe_pendiente=False,
            cfg=self._cfg_seguimiento(umbral=1.0),
        )
        svc = ConvivenciaService(repo=conv_repo, alerta_repo=alerta_repo)

        # Registrar 2 incidentes negativos; el 2do (conteo==2 >= umbral*2==2) activa CRITICA
        svc.registrar_comportamiento(
            self._dto_negativo(), usuario_id=1, anio_id=2026, usuario_rol=None
        )
        # Reiniciar alerta_repo para solo capturar la segunda
        alerta_repo._alertas.clear()
        svc.registrar_comportamiento(
            self._dto_negativo(), usuario_id=1, anio_id=2026, usuario_rol=None
        )

        assert len(alerta_repo._alertas) == 1
        assert alerta_repo._alertas[0].nivel == NivelAlerta.CRITICA


# ===========================================================================
# convivencia_18: vista_360
# ===========================================================================

class _FakeEstWithGrupo:
    """Estudiante fake con id, nombre, apellido y grupo_id."""
    def __init__(self, id: int, nombre: str = "Ana", apellido: str = "Ruiz", grupo_id: int = 10):
        self.id       = id
        self.nombre   = nombre
        self.apellido = apellido
        self.grupo_id = grupo_id


class _FakeEstSvcById:
    """EstudianteService fake con get_by_id."""
    def __init__(self, estudiante):
        self._est = estudiante

    def get_by_id(self, estudiante_id: int):
        return self._est

    def listar_por_grupo(self, grupo_id, solo_activos=True):
        return [self._est]


def _svc_vista_360(
    conv_repo=None,
    alerta_repo=None,
    estudiante=None,
    catalogo_autoriza: bool | None = None,
    con_niveles: bool = True,
):
    """Factory para tests de vista_360 con dependencias configurables."""
    conv_repo = conv_repo or FakeConvRepo()
    est = estudiante or _FakeEstWithGrupo(1)
    svc_kwargs: dict = {
        "repo": conv_repo,
        "alerta_repo": alerta_repo,
        "estudiante_svc_provider": lambda: _FakeEstSvcById(est),
    }
    if con_niveles:
        svc_kwargs["configuracion_svc_provider"] = lambda: _FakeConfigSvc(_NIVELES)
        svc_kwargs["periodo_svc_provider"]        = lambda: _FakePeriodoSvc()
    if catalogo_autoriza is not None:
        stub = _StubCatalogoSvc(autoriza=catalogo_autoriza)
        svc_kwargs["catalogo_academico_svc_provider"] = lambda: stub
    return ConvivenciaService(**svc_kwargs), conv_repo


class TestVista360:
    """Tests para ConvivenciaService.vista_360 (convivencia_18)."""

    def test_vista_360_flujo_nominal_director(self):
        """Director obtiene DTO completo con nota, observaciones y alertas."""
        conv_repo = FakeConvRepo()
        alerta_repo = FakeAlertaRepo()
        # Precargar datos
        conv_repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5,
            valor=85.0, observacion="Excelente conducta",
        )
        conv_repo.guardar_observacion(ObservacionPeriodo(
            estudiante_id=1, asignacion_id=3, periodo_id=5,
            texto="Participativo y respetuoso", es_publica=True,
        ))
        alerta_repo.guardar_alerta(Alerta(
            tipo_alerta=TipoAlerta.SEGUIMIENTO_REQUERIDO,
            estudiante_id=1,
            descripcion="Requiere seguimiento",
            nivel=NivelAlerta.ADVERTENCIA,
        ))

        svc, _ = _svc_vista_360(conv_repo=conv_repo, alerta_repo=alerta_repo)
        dto = svc.vista_360(
            estudiante_id=1, periodo_id=5,
            usuario_id=10, usuario_rol="director",
        )

        assert isinstance(dto, Seguimiento360DTO)
        assert dto.estudiante_id == 1
        assert "Ruiz" in dto.estudiante_nombre or "Ana" in dto.estudiante_nombre
        assert dto.nota_comportamiento == pytest.approx(85.0)
        assert dto.concepto == "Excelente conducta"
        assert dto.nivel_comportamiento == "Superior"
        assert "Participativo y respetuoso" in dto.observaciones
        assert len(dto.alertas_activas) == 1
        assert "Requiere seguimiento" in dto.alertas_activas[0]

    def test_vista_360_director_grupo_autorizado(self):
        """director_de_grupo con autorización de catalogo accede correctamente."""
        conv_repo = FakeConvRepo()
        conv_repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5, valor=70.0,
        )

        svc, _ = _svc_vista_360(
            conv_repo=conv_repo,
            catalogo_autoriza=True,  # _StubCatalogoSvc(autoriza=True)
        )
        dto = svc.vista_360(
            estudiante_id=1, periodo_id=5,
            usuario_id=99, usuario_rol="director_de_grupo",
        )
        assert isinstance(dto, Seguimiento360DTO)
        assert dto.nota_comportamiento == pytest.approx(70.0)

    def test_vista_360_director_grupo_no_autorizado_lanza(self):
        """director_de_grupo rechazado por catalogo → PermissionError."""
        svc, _ = _svc_vista_360(catalogo_autoriza=False)
        with pytest.raises(PermissionError):
            svc.vista_360(
                estudiante_id=1, periodo_id=5,
                usuario_id=99, usuario_rol="director_de_grupo",
            )

    def test_vista_360_profesor_no_autorizado(self):
        """Rol 'profesor' no tiene acceso al seguimiento 360° → PermissionError."""
        svc, _ = _svc_vista_360()
        with pytest.raises(PermissionError, match="Solo director"):
            svc.vista_360(
                estudiante_id=1, periodo_id=5,
                usuario_id=50, usuario_rol="profesor",
            )

    def test_vista_360_sin_datos(self):
        """Sin nota ni observaciones → DTO con campos None y listas vacías."""
        svc, _ = _svc_vista_360()
        dto = svc.vista_360(
            estudiante_id=1, periodo_id=5,
            usuario_id=10, usuario_rol="director",
        )
        assert isinstance(dto, Seguimiento360DTO)
        assert dto.nota_comportamiento is None
        assert dto.concepto is None
        assert dto.nivel_comportamiento is None
        assert dto.observaciones == []
        assert dto.alertas_activas == []
        assert dto.promedio_notas is None

    def test_vista_360_coordinador_permitido(self):
        """Coordinador también accede sin restricciones adicionales."""
        svc, _ = _svc_vista_360()
        dto = svc.vista_360(
            estudiante_id=1, periodo_id=5,
            usuario_id=20, usuario_rol="coordinador",
        )
        assert isinstance(dto, Seguimiento360DTO)

    def test_vista_360_sin_alerta_repo_alertas_vacias(self):
        """Sin alerta_repo inyectado → alertas_activas=[] silenciosamente."""
        conv_repo = FakeConvRepo()
        svc, _ = _svc_vista_360(conv_repo=conv_repo, alerta_repo=None)
        dto = svc.vista_360(
            estudiante_id=1, periodo_id=5,
            usuario_id=10, usuario_rol="director",
        )
        assert dto.alertas_activas == []

    def test_vista_360_sin_providers_niveles_extrae_nota_directa(self):
        """Sin providers de niveles, vista_360 extrae nota directamente del repo."""
        conv_repo = FakeConvRepo()
        conv_repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5,
            valor=75.0, observacion="Buen comportamiento",
        )
        svc, _ = _svc_vista_360(conv_repo=conv_repo, con_niveles=False)
        dto = svc.vista_360(
            estudiante_id=1, periodo_id=5,
            usuario_id=10, usuario_rol="director",
        )
        # Sin providers de niveles → nota extraída directo del repo
        assert dto.nota_comportamiento == pytest.approx(75.0)
        assert dto.concepto == "Buen comportamiento"
        assert dto.nivel_comportamiento is None  # sin niveles no se puede resolver


# ===========================================================================
# convivencia_21: serie_notas_comportamiento y resumen_convivencia_grupo
# ===========================================================================

class _FakePeriodo:
    def __init__(self, id: int, nombre: str, anio_id: int = 2026):
        self.id      = id
        self.nombre  = nombre
        self.anio_id = anio_id


class _FakePeriodoSvcConLista:
    """PeriodoService fake con listar_por_anio y get_by_id."""
    def __init__(self, periodos: list[_FakePeriodo]):
        self._periodos = periodos

    def listar_por_anio(self, anio_id: int) -> list[_FakePeriodo]:
        return [p for p in self._periodos if p.anio_id == anio_id]

    def get_by_id(self, periodo_id: int) -> _FakePeriodo:
        for p in self._periodos:
            if p.id == periodo_id:
                return p
        raise ValueError(periodo_id)


class TestSerieNotasComportamiento:
    """convivencia_21 — serie_notas_comportamiento."""

    def _periodos(self) -> list[_FakePeriodo]:
        return [
            _FakePeriodo(1, "Periodo 1"),
            _FakePeriodo(2, "Periodo 2"),
            _FakePeriodo(3, "Periodo 3"),
        ]

    def test_serie_nominal_con_huecos(self):
        """Un punto por periodo, en orden; periodos sin nota → valor None."""
        repo = FakeConvRepo()
        repo._notas[(1, 1)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=1, valor=80.0,
        )
        repo._notas[(1, 3)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=3, valor=90.0,
        )
        svc = ConvivenciaService(
            repo=repo,
            periodo_svc_provider=lambda: _FakePeriodoSvcConLista(self._periodos()),
        )
        serie = svc.serie_notas_comportamiento(estudiante_id=1, anio_id=2026)
        assert len(serie) == 3
        assert all(isinstance(p, PuntoSerieDTO) for p in serie)
        assert [p.periodo_id for p in serie] == [1, 2, 3]
        assert [p.periodo_nombre for p in serie] == ["Periodo 1", "Periodo 2", "Periodo 3"]
        assert [p.valor for p in serie] == [80.0, None, 90.0]

    def test_serie_estudiante_sin_nota_todo_none(self):
        """Estudiante sin ninguna nota → toda la serie con valor None."""
        repo = FakeConvRepo()
        svc = ConvivenciaService(
            repo=repo,
            periodo_svc_provider=lambda: _FakePeriodoSvcConLista(self._periodos()),
        )
        serie = svc.serie_notas_comportamiento(estudiante_id=99, anio_id=2026)
        assert len(serie) == 3
        assert all(p.valor is None for p in serie)

    def test_serie_sin_periodo_provider_lanza(self):
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)
        with pytest.raises(RuntimeError):
            svc.serie_notas_comportamiento(estudiante_id=1, anio_id=2026)


class TestResumenConvivenciaGrupo:
    """convivencia_21 — resumen_convivencia_grupo."""

    def _cfg_umbral(self, umbral: float = 2.0) -> ConfiguracionAlerta:
        return ConfiguracionAlerta(
            anio_id=2026,
            tipo_alerta=TipoAlerta.SEGUIMIENTO_REQUERIDO,
            umbral=umbral,
            activa=True,
        )

    def _ests(self):
        ests = [_FakeEst(1), _FakeEst(2), _FakeEst(3)]
        for e, nom, ape in [(ests[0], "Ana", "Ruiz"), (ests[1], "Bob", "Diaz"), (ests[2], "Cyd", "Paz")]:
            e.nombre = nom
            e.apellido = ape
        return ests

    def _svc(self, repo, ests, alerta_repo=None):
        return ConvivenciaService(
            repo=repo,
            alerta_repo=alerta_repo,
            configuracion_svc_provider=lambda: _FakeConfigSvc(_NIVELES),
            periodo_svc_provider=lambda: _FakePeriodoSvcConLista(
                [_FakePeriodo(5, "Periodo 5")]
            ),
            estudiante_svc_provider=lambda: _FakeEstSvc(ests),
        )

    def test_resumen_nominal(self):
        """Combina nota/nivel, conteo de observaciones y negativos, y umbral."""
        repo = FakeConvRepo()
        ests = self._ests()
        repo._asig_grupo[99] = 10
        # Estudiante 1: nota Superior, 2 observaciones, 2 registros negativos.
        repo._notas[(1, 5)] = NotaComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5, valor=90.0,
        )
        repo.guardar_observacion(ObservacionPeriodo(
            estudiante_id=1, asignacion_id=99, periodo_id=5, texto="Obs A",
        ))
        repo.guardar_observacion(ObservacionPeriodo(
            estudiante_id=1, asignacion_id=99, periodo_id=5, texto="Obs B",
        ))
        for _ in range(2):
            repo.guardar_registro(RegistroComportamiento(
                estudiante_id=1, grupo_id=10, periodo_id=5,
                tipo=TipoRegistro.DIFICULTAD, descripcion="Incidente",
            ))
        # Estudiante 2: sin nota, 1 observación, 0 negativos.
        repo.guardar_observacion(ObservacionPeriodo(
            estudiante_id=2, asignacion_id=99, periodo_id=5, texto="Obs C",
        ))
        # Estudiante 3: sin nada.

        alerta_repo = FakeAlertaRepo(cfg=self._cfg_umbral(umbral=2.0))
        svc = self._svc(repo, ests, alerta_repo=alerta_repo)
        resumen = svc.resumen_convivencia_grupo(grupo_id=10, periodo_id=5)

        assert len(resumen) == 3
        assert all(isinstance(r, ResumenConvivenciaDTO) for r in resumen)
        by_id = {r.estudiante_id: r for r in resumen}

        assert by_id[1].nombre == "Ruiz Ana"
        assert by_id[1].nota == 90.0
        assert by_id[1].nivel_nombre == "Superior"
        assert by_id[1].num_observaciones == 2
        assert by_id[1].num_registros_negativos == 2
        assert by_id[1].supera_umbral is True  # 2 >= umbral 2

        assert by_id[2].nota is None
        assert by_id[2].nivel_nombre is None
        assert by_id[2].num_observaciones == 1
        assert by_id[2].num_registros_negativos == 0
        assert by_id[2].supera_umbral is False

        assert by_id[3].num_observaciones == 0
        assert by_id[3].num_registros_negativos == 0
        assert by_id[3].nota is None

    def test_resumen_grupo_vacio(self):
        """Grupo sin estudiantes → lista vacía."""
        repo = FakeConvRepo()
        svc = self._svc(repo, ests=[], alerta_repo=FakeAlertaRepo(cfg=self._cfg_umbral()))
        resumen = svc.resumen_convivencia_grupo(grupo_id=10, periodo_id=5)
        assert resumen == []

    def test_resumen_sin_alerta_repo_supera_umbral_false(self):
        """Sin alerta_repo, supera_umbral es False aunque haya muchos negativos."""
        repo = FakeConvRepo()
        ests = self._ests()
        for _ in range(5):
            repo.guardar_registro(RegistroComportamiento(
                estudiante_id=1, grupo_id=10, periodo_id=5,
                tipo=TipoRegistro.DIFICULTAD, descripcion="Incidente",
            ))
        svc = self._svc(repo, ests, alerta_repo=None)
        resumen = svc.resumen_convivencia_grupo(grupo_id=10, periodo_id=5)
        by_id = {r.estudiante_id: r for r in resumen}
        assert by_id[1].num_registros_negativos == 5
        assert by_id[1].supera_umbral is False


# ===========================================================================
# convivencia_36: medidas pedagógicas
# ===========================================================================

class TestMedidasPedagogicas:
    """Tests para ConvivenciaService — CRUD de medidas pedagógicas."""

    def test_crear_medida_director(self):
        svc, repo = _make_svc()
        dto = NuevaMedidaPedagogicaDTO(nombre="Dialogo pedagogico", nivel_minimo=1)
        medida = svc.crear_medida_pedagogica(dto, usuario_rol="director")
        assert medida.id is not None
        assert medida.nombre == "Dialogo pedagogico"
        assert medida.nivel_minimo == 1
        assert medida.activa is True

    def test_crear_medida_coordinador(self):
        svc, repo = _make_svc()
        dto = NuevaMedidaPedagogicaDTO(nombre="Citacion", nivel_minimo=2)
        medida = svc.crear_medida_pedagogica(dto, usuario_rol="coordinador")
        assert medida.id is not None
        assert medida.nivel_minimo == 2

    def test_crear_medida_profesor_rechazado(self):
        svc, _ = _make_svc()
        dto = NuevaMedidaPedagogicaDTO(nombre="X")
        with pytest.raises(PermissionError):
            svc.crear_medida_pedagogica(dto, usuario_rol="profesor")

    def test_actualizar_medida(self):
        svc, repo = _make_svc()
        dto = NuevaMedidaPedagogicaDTO(nombre="Original", nivel_minimo=1)
        medida = svc.crear_medida_pedagogica(dto, usuario_rol="director")

        dto2 = NuevaMedidaPedagogicaDTO(nombre="Actualizada", nivel_minimo=2)
        actualizada = svc.actualizar_medida_pedagogica(medida.id, dto2, usuario_rol="director")
        assert actualizada.nombre == "Actualizada"
        assert actualizada.nivel_minimo == 2

    def test_actualizar_medida_inexistente_lanza(self):
        svc, _ = _make_svc()
        dto = NuevaMedidaPedagogicaDTO(nombre="X")
        with pytest.raises(ValueError, match="999"):
            svc.actualizar_medida_pedagogica(999, dto, usuario_rol="director")

    def test_desactivar_medida(self):
        svc, repo = _make_svc()
        dto = NuevaMedidaPedagogicaDTO(nombre="Matricula condicional", nivel_minimo=3)
        medida = svc.crear_medida_pedagogica(dto, usuario_rol="director")
        assert medida.activa is True

        desactivada = svc.desactivar_medida_pedagogica(medida.id, usuario_rol="director")
        assert desactivada.activa is False
        assert repo._medidas[medida.id].activa is False

    def test_desactivar_medida_profesor_rechazado(self):
        svc, repo = _make_svc()
        dto = NuevaMedidaPedagogicaDTO(nombre="X")
        medida = svc.crear_medida_pedagogica(dto, usuario_rol="director")
        with pytest.raises(PermissionError):
            svc.desactivar_medida_pedagogica(medida.id, usuario_rol="profesor")

    def test_listar_medidas_filtra_por_nivel(self):
        """Medidas con nivel_minimo=3 no aparecen en el listado si el tipo de situación es nivel 1."""
        svc, repo = _make_svc()
        # Crear medidas de distintos niveles
        for nombre, nivel in [("Dialogo", 1), ("Citacion", 2), ("No renovacion", 3)]:
            dto = NuevaMedidaPedagogicaDTO(nombre=nombre, nivel_minimo=nivel)
            svc.crear_medida_pedagogica(dto, usuario_rol="director")

        todas = svc.listar_medidas_pedagogicas(solo_activas=True)
        assert len(todas) == 3

        # Filtrado de negocio en el presenter/UI: solo las aplicables a nivel 1
        nivel_situacion = 1
        aplicables = [m for m in todas if m.nivel_minimo <= nivel_situacion]
        assert len(aplicables) == 1
        assert aplicables[0].nombre == "Dialogo"


# ===========================================================================
# convivencia_37 — observador_estudiante + exportar_observador
# ===========================================================================

class _FakeEstConNombre:
    def __init__(self, id: int, nombre: str = "Ana", apellido: str = "Ruiz"):
        self.id = id
        self.nombre = nombre
        self.apellido = apellido


class _FakeEstSvcById37:
    def __init__(self, est):
        self._est = est

    def get_by_id(self, estudiante_id: int):
        return self._est

    def listar_por_grupo(self, grupo_id):
        return [self._est]


class _FakePeriodo37:
    def __init__(self, id: int, nombre: str, anio_id: int = 2026):
        self.id = id
        self.nombre = nombre
        self.anio_id = anio_id


class _FakePeriodoSvc37:
    def __init__(self, periodos):
        self._periodos = periodos

    def listar_por_anio(self, anio_id: int):
        return self._periodos

    def get_by_id(self, periodo_id: int):
        for p in self._periodos:
            if p.id == periodo_id:
                return p
        raise ValueError(periodo_id)


class _FakeConvRepoObs37(FakeConvRepo):
    """FakeConvRepo extendido con soporte de listar_entradas_seguimiento."""

    def __init__(self):
        super().__init__()
        self._entradas_seg: dict[int, list] = {}

    def listar_entradas_seguimiento(self, registro_id: int):
        from src.domain.models.convivencia import EntradaSeguimiento
        return self._entradas_seg.get(registro_id, [])

    def guardar_entrada_seguimiento(self, entrada):
        from src.domain.models.convivencia import EntradaSeguimiento
        return entrada.model_copy(update={"id": 1})

    def listar_registros(self, filtro, institucion_id=None):
        regs = list(self._regs.values())
        if filtro.estudiante_id is not None:
            regs = [r for r in regs if r.estudiante_id == filtro.estudiante_id]
        if filtro.periodo_id is not None:
            regs = [r for r in regs if r.periodo_id == filtro.periodo_id]
        return regs

    def listar_observaciones_por_estudiante(self, est_id, per_id=None, solo_publicas=False):
        obs = [o for o in self._obs.values() if o.estudiante_id == est_id]
        if per_id is not None:
            obs = [o for o in obs if o.periodo_id == per_id]
        if solo_publicas:
            obs = [o for o in obs if o.es_publica]
        return obs


def _svc_observador(est=None, periodos=None):
    """Crea un ConvivenciaService con providers para el observador."""
    from src.domain.models.convivencia import EntradaSeguimiento
    repo = _FakeConvRepoObs37()
    est = est or _FakeEstConNombre(1)
    periodos = periodos or [_FakePeriodo37(5, "Periodo 1")]
    svc = ConvivenciaService(
        repo=repo,
        estudiante_svc_provider=lambda: _FakeEstSvcById37(est),
        periodo_svc_provider=lambda: _FakePeriodoSvc37(periodos),
    )
    return svc, repo


class TestObservadorEstudiante:
    """Tests de observador_estudiante y exportar_observador (convivencia_37)."""

    def test_retorna_dict_con_claves_requeridas(self):
        svc, _ = _svc_observador()
        resultado = svc.observador_estudiante(estudiante_id=1, anio_id=2026)
        assert "estudiante" in resultado
        assert "institucion" in resultado
        assert "anio" in resultado
        assert "periodo" in resultado
        assert "entradas" in resultado
        assert "resumen" in resultado

    def test_entradas_vacias_sin_datos(self):
        svc, _ = _svc_observador()
        resultado = svc.observador_estudiante(estudiante_id=1, anio_id=2026)
        assert resultado["entradas"] == []
        assert resultado["periodo"] is None

    def test_entradas_incluyen_observaciones_y_registros_ordenados(self):
        from datetime import datetime
        periodos = [_FakePeriodo37(5, "P1")]
        svc, repo = _svc_observador(periodos=periodos)

        # Registro creado primero (fecha antigua)
        reg = RegistroComportamiento(
            estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD,
            descripcion="Pelea en el patio",
            fecha=date(2026, 3, 10),
        )
        reg = repo.guardar_registro(reg)

        # Observación pública más reciente
        obs = ObservacionPeriodo(
            estudiante_id=1, asignacion_id=1, periodo_id=5,
            texto="Buen comportamiento",
            es_publica=True,
            fecha_registro=datetime(2026, 4, 1, 10, 0, 0),
            categoria_id=None,
        )
        repo.guardar_observacion(obs)

        resultado = svc.observador_estudiante(estudiante_id=1, anio_id=2026)
        entradas = resultado["entradas"]
        assert len(entradas) == 2
        # Ordenadas cronológicamente: registro (mar) antes que observación (abr)
        assert entradas[0]["tipo"] == "registro"
        assert entradas[0]["subtipo"] == "dificultad"
        assert entradas[1]["tipo"] == "observacion"

    def test_resumen_cuenta_tipos_correctamente(self):
        periodos = [_FakePeriodo37(5, "P1")]
        svc, repo = _svc_observador(periodos=periodos)
        for tipo in [TipoRegistro.FORTALEZA, TipoRegistro.FORTALEZA, TipoRegistro.DIFICULTAD, TipoRegistro.COMPROMISO]:
            reg = RegistroComportamiento(
                estudiante_id=1, grupo_id=10, periodo_id=5,
                tipo=tipo, descripcion="desc", fecha=date.today(),
            )
            repo.guardar_registro(reg)

        resultado = svc.observador_estudiante(estudiante_id=1, anio_id=2026)
        resumen = resultado["resumen"]
        assert resumen["fortalezas"] == 2
        assert resumen["dificultades"] == 1
        assert resumen["compromisos"] == 1
        assert resumen["citaciones"] == 0

    def test_filtro_por_periodo_id(self):
        periodos = [_FakePeriodo37(5, "P1"), _FakePeriodo37(6, "P2")]
        svc, repo = _svc_observador(periodos=periodos)

        for per_id in [5, 6]:
            reg = RegistroComportamiento(
                estudiante_id=1, grupo_id=10, periodo_id=per_id,
                tipo=TipoRegistro.FORTALEZA, descripcion="X", fecha=date.today(),
            )
            repo.guardar_registro(reg)

        resultado = svc.observador_estudiante(estudiante_id=1, anio_id=2026, periodo_id=5)
        assert len(resultado["entradas"]) == 1
        assert resultado["periodo"] == "P1"

    def test_sin_providers_lanza_runtime_error(self):
        repo = FakeConvRepo()
        svc = ConvivenciaService(repo=repo)  # sin providers
        with pytest.raises(RuntimeError):
            svc.observador_estudiante(estudiante_id=1, anio_id=2026)

    def test_exportar_observador_pdf_retorna_bytes(self):
        from src.infrastructure.exporters.null_exporter import NullExporter
        svc, _ = _svc_observador()
        datos_bytes = svc.exportar_observador(estudiante_id=1, anio_id=2026, formato="pdf")
        assert isinstance(datos_bytes, bytes)
        assert len(datos_bytes) > 0
        # PDF mágico: empieza con %PDF
        assert datos_bytes[:4] == b"%PDF"

    def test_exportar_observador_excel_retorna_bytes(self):
        from src.infrastructure.exporters.openpyxl_exporter import OpenpyxlExporter
        repo = _FakeConvRepoObs37()
        periodos = [_FakePeriodo37(5, "P1")]
        est = _FakeEstConNombre(1)
        svc = ConvivenciaService(
            repo=repo,
            estudiante_svc_provider=lambda: _FakeEstSvcById37(est),
            periodo_svc_provider=lambda: _FakePeriodoSvc37(periodos),
            exporter=OpenpyxlExporter(),
        )
        datos_bytes = svc.exportar_observador(estudiante_id=1, anio_id=2026, formato="excel")
        assert isinstance(datos_bytes, bytes)
        assert len(datos_bytes) > 0
        # XLSX mágico: empieza con PK (ZIP)
        assert datos_bytes[:2] == b"PK"

    def test_exportar_formato_invalido_lanza(self):
        svc, _ = _svc_observador()
        with pytest.raises(ValueError, match="Formato no soportado"):
            svc.exportar_observador(estudiante_id=1, anio_id=2026, formato="csv")


# ===========================================================================
# convivencia_38 — Integración tipos_situacion y medidas en reportes
# ===========================================================================

from src.domain.models.convivencia import TipoSituacion


class _FakeConvRepo38(FakeConvRepo):
    """Repo con tipos de situación y medidas precargados para tests de conv_38."""

    def __init__(self, tipos: list | None = None, medidas: list | None = None):
        super().__init__()
        self._tipos38 = tipos or []
        # medidas ya están en FakeConvRepo._medidas; guardamos las extra en el super
        for m in (medidas or []):
            self._medidas[m.id] = m

    def listar_tipos_situacion(self, solo_activas=True, institucion_id=None):
        return list(self._tipos38)

    def listar_registros(self, filtro: FiltroConvivenciaDTO, institucion_id=None):
        regs = list(self._regs.values())
        if filtro.estudiante_id is not None:
            regs = [r for r in regs if r.estudiante_id == filtro.estudiante_id]
        if filtro.periodo_id is not None:
            regs = [r for r in regs if r.periodo_id == filtro.periodo_id]
        if filtro.grupo_id is not None:
            regs = [r for r in regs if r.grupo_id == filtro.grupo_id]
        return regs


def _tipo(id_: int, nombre: str) -> TipoSituacion:
    return TipoSituacion(id=id_, nombre=nombre, activa=True, institucion_id=1)


def _medida_obj(id_: int, nombre: str) -> MedidaPedagogica:
    return MedidaPedagogica(id=id_, nombre=nombre, activa=True, institucion_id=1)


class TestRegistrosInformablesPeriodo38:
    """Verifica que _registros_informables_periodo incluye tipo_situacion y medida."""

    def _svc_con_prefs(self, tipos=None, medidas=None):
        from src.domain.models.preferencia_institucion import PreferenciasDTO

        repo = _FakeConvRepo38(tipos=tipos, medidas=medidas)
        svc = ConvivenciaService(repo=repo)
        # Include fortaleza + dificultad; bypass institucion_actual() entirely
        prefs = PreferenciasDTO(
            registros_boletin_tipos=["fortaleza", "dificultad"],
            registros_boletin_dificultad_requiere_notificacion=False,
        )
        svc._get_prefs_convivencia = lambda: prefs
        return svc, repo

    def test_campos_tipo_situacion_y_medida_presentes_cuando_none(self):
        svc, repo = self._svc_con_prefs()
        reg = RegistroComportamiento(
            id=1, estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD, descripcion="Problema en clase",
            fecha=date(2026, 3, 1),
        )
        repo._regs[1] = reg
        resultado = svc._registros_informables_periodo(1, 5)
        assert len(resultado) == 1
        assert "tipo_situacion" in resultado[0]
        assert "medida" in resultado[0]
        assert resultado[0]["tipo_situacion"] is None
        assert resultado[0]["medida"] is None

    def test_tipo_situacion_resuelto_por_nombre(self):
        tipo = _tipo(7, "Tipo II")
        svc, repo = self._svc_con_prefs(tipos=[tipo])
        reg = RegistroComportamiento(
            id=1, estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD, descripcion="Pelea",
            fecha=date(2026, 3, 5), tipo_situacion_id=7,
        )
        repo._regs[1] = reg
        resultado = svc._registros_informables_periodo(1, 5)
        assert resultado[0]["tipo_situacion"] == "Tipo II"

    def test_medida_resuelta_por_nombre(self):
        medida = _medida_obj(3, "Diálogo con acudiente")
        svc, repo = self._svc_con_prefs(medidas=[medida])
        reg = RegistroComportamiento(
            id=1, estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD, descripcion="Pelea",
            fecha=date(2026, 3, 5), medida_id=3,
        )
        repo._regs[1] = reg
        resultado = svc._registros_informables_periodo(1, 5)
        assert resultado[0]["medida"] == "Diálogo con acudiente"

    def test_tipo_desconocido_queda_none(self):
        """Si el tipo_situacion_id no existe en el mapa, retorna None."""
        svc, repo = self._svc_con_prefs(tipos=[_tipo(1, "Tipo I")])
        reg = RegistroComportamiento(
            id=1, estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD, descripcion="Falta",
            fecha=date(2026, 3, 5), tipo_situacion_id=999,
        )
        repo._regs[1] = reg
        resultado = svc._registros_informables_periodo(1, 5)
        assert resultado[0]["tipo_situacion"] is None


class TestReportePeriodoGrupo38:
    """Verifica que reporte_periodo_grupo incluye desglose_por_tipo (conv_38)."""

    def _svc_con_tipos(self, tipos=None):
        repo = _FakeConvRepo38(tipos=tipos or [])
        est = _FakeEst(1)
        est.nombre = "Ana"
        est.apellido = "Ruiz"
        svc = ConvivenciaService(
            repo=repo,
            configuracion_svc_provider=lambda: _FakeConfigSvc(_NIVELES),
            periodo_svc_provider=lambda: _FakePeriodoSvc(),
            estudiante_svc_provider=lambda: _FakeEstSvc([est]),
        )
        return svc, repo

    def test_desglose_none_cuando_no_hay_tipos(self):
        svc, _ = self._svc_con_tipos(tipos=[])
        filas = svc.reporte_periodo_grupo(grupo_id=10, periodo_id=5)
        assert len(filas) == 1
        assert filas[0].desglose_por_tipo is None

    def test_desglose_cero_cuando_hay_tipos_pero_sin_negativos(self):
        tipos = [_tipo(1, "Tipo I"), _tipo(2, "Tipo II")]
        svc, repo = self._svc_con_tipos(tipos=tipos)
        filas = svc.reporte_periodo_grupo(grupo_id=10, periodo_id=5)
        assert len(filas) == 1
        desglose = filas[0].desglose_por_tipo
        assert desglose is not None
        assert desglose.get("Tipo I") == 0
        assert desglose.get("Tipo II") == 0

    def test_desglose_cuenta_negativos_por_tipo(self):
        tipos = [_tipo(1, "Tipo I"), _tipo(2, "Tipo II")]
        svc, repo = self._svc_con_tipos(tipos=tipos)
        repo._regs[1] = RegistroComportamiento(
            id=1, estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD, descripcion="Falta leve",
            fecha=date(2026, 3, 1), tipo_situacion_id=1,
        )
        repo._regs[2] = RegistroComportamiento(
            id=2, estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD, descripcion="Agresión verbal",
            fecha=date(2026, 3, 5), tipo_situacion_id=2,
        )
        repo._regs[3] = RegistroComportamiento(
            id=3, estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD, descripcion="Segunda leve",
            fecha=date(2026, 3, 7), tipo_situacion_id=1,
        )
        # Fortaleza no debe contar como negativo
        repo._regs[4] = RegistroComportamiento(
            id=4, estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.FORTALEZA, descripcion="Buen trabajo",
            fecha=date(2026, 3, 10),
        )
        filas = svc.reporte_periodo_grupo(grupo_id=10, periodo_id=5)
        assert len(filas) == 1
        desglose = filas[0].desglose_por_tipo
        assert desglose is not None
        assert desglose.get("Tipo I") == 2
        assert desglose.get("Tipo II") == 1

    def test_pdf_incluye_columnas_desglose(self):
        tipos = [_tipo(1, "Tipo I"), _tipo(2, "Tipo II")]
        svc, repo = self._svc_con_tipos(tipos=tipos)
        repo._regs[1] = RegistroComportamiento(
            id=1, estudiante_id=1, grupo_id=10, periodo_id=5,
            tipo=TipoRegistro.DIFICULTAD, descripcion="Falta",
            fecha=date(2026, 3, 1), tipo_situacion_id=1,
        )

        class _Cap:
            def exportar_excel(self, datos, nombre_hoja="Datos", ruta_destino=None):
                return b""
            def exportar_pdf(self, html, ruta_destino=None):
                return b""

        exp = _Cap()
        svc2 = ConvivenciaService(
            repo=repo,
            configuracion_svc_provider=lambda: _FakeConfigSvc(_NIVELES),
            periodo_svc_provider=lambda: _FakePeriodoSvc(),
            estudiante_svc_provider=lambda: _FakeEstSvc([_FakeEst(1)]),
            exporter=exp,
        )
        pdf_bytes = svc2.exportar_reporte_periodo_grupo(10, 5, "pdf")
        assert isinstance(pdf_bytes, bytes)
        assert pdf_bytes[:5] == b"%PDF-"



"""
observador_excel.py — Generador del Observador del Estudiante en Excel.

Espeja la estructura del PDF (observador_pdf.py):
  1. Membrete institucional
  2. Ficha de identificación del estudiante
  3. Tabla de observaciones académicas
  4. Tabla de registros de comportamiento
  5. Detalle de seguimientos
  6. Resumen estadístico
  7. Notas de comportamiento por periodo
"""

from __future__ import annotations

import io
from datetime import date as _date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Estilos ─────────────────────────────────────────────────────────────────

_AZUL_OSCURO = "1A365D"
_AZUL = "2B6CB0"
_AZUL_CLARO = "DBEAFE"
_GRIS_FONDO = "F7FAFC"
_BLANCO = "FFFFFF"
_VERDE = "276749"
_ROJO = "9B2C2C"
_AMARILLO = "975A16"
_NARANJA = "C05621"
_MORADO = "6B46C1"
_GRIS_LINEA = "CBD5E0"

_HEADER_FILL = PatternFill(fill_type="solid", fgColor=_AZUL)
_HEADER_FONT = Font(bold=True, color=_BLANCO, size=9)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_SECCION_FILL = PatternFill(fill_type="solid", fgColor=_AZUL_OSCURO)
_SECCION_FONT = Font(bold=True, color=_BLANCO, size=10)

_LABEL_FILL = PatternFill(fill_type="solid", fgColor=_AZUL_CLARO)
_LABEL_FONT = Font(bold=True, size=9, color="374151")

_VALUE_FONT = Font(size=9, color="374151")
_VALUE_ALIGN = Alignment(vertical="center", wrap_text=True)

_TITULO_FONT = Font(bold=True, size=13, color=_AZUL_OSCURO)
_TITULO_ALIGN = Alignment(horizontal="center", vertical="center")

_INST_FONT = Font(bold=True, size=12, color=_AZUL_OSCURO)
_INST_ALIGN = Alignment(horizontal="center", vertical="center")
_INST_SUB_FONT = Font(size=8, color="718096")
_INST_SUB_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THIN_BORDER = Border(
    left=Side(style="thin", color=_GRIS_LINEA),
    right=Side(style="thin", color=_GRIS_LINEA),
    top=Side(style="thin", color=_GRIS_LINEA),
    bottom=Side(style="thin", color=_GRIS_LINEA),
)

_SEG_HEADER_FILL = PatternFill(fill_type="solid", fgColor="4A5568")

_TIPO_LABEL = {
    "fortaleza": "Fortaleza",
    "dificultad": "Dificultad",
    "compromiso": "Compromiso",
    "citacion_acudiente": "Citación acudiente",
    "descargo": "Descargo",
}

_RESUMEN_COLORES = {
    "Observaciones": _AZUL,
    "Fortalezas": _VERDE,
    "Dificultades": _ROJO,
    "Compromisos": _AMARILLO,
    "Citaciones": _NARANJA,
    "Descargos": _MORADO,
}


# ── Helpers ─────────────────────────────────────────────────────────────────


def _fecha_corta(dt) -> str:
    if dt is None:
        return "—"
    try:
        return f"{dt.day:02d}/{dt.month:02d}/{dt.year}"
    except Exception:
        return str(dt)


def _fecha_hora(dt) -> str:
    if dt is None:
        return "—"
    try:
        base = f"{dt.day:02d}/{dt.month:02d}/{dt.year}"
        if hasattr(dt, "hour") and hasattr(dt, "minute"):
            return f"{base} {dt.hour:02d}:{dt.minute:02d}"
        return base
    except Exception:
        return str(dt)


def _safe(val) -> str:
    if val is None or str(val).strip() == "":
        return "—"
    return str(val)


def _write_row(ws, row: int, values: list, font=None, fill=None, alignment=None, border=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row, col, _safe(val) if isinstance(val, str) else val)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border


def _apply_borders(ws, start_row: int, end_row: int, num_cols: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            ws.cell(r, c).border = _THIN_BORDER


def _set_col_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _zebra_fill(row_idx: int) -> PatternFill | None:
    if row_idx % 2 == 0:
        return PatternFill(fill_type="solid", fgColor=_GRIS_FONDO)
    return None


# ── Función principal ───────────────────────────────────────────────────────


def generar_observador_excel(datos: dict) -> bytes:
    """Genera el Excel del observador del estudiante y retorna los bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active

    est = datos.get("estudiante", {})
    inst = datos.get("institucion", {})
    anio = str(datos.get("anio", ""))
    periodo = datos.get("periodo")
    entradas = datos.get("entradas", [])
    resumen = datos.get("resumen", {})

    nombre_est = est.get("nombre", "Estudiante")
    ws.title = f"Observador_{nombre_est}"[:31]

    _set_col_widths(ws, [5, 14, 12, 12, 12, 50, 14, 7, 14])

    row = 1

    # ── 1. Membrete institucional ───────────────────────────────────────
    nombre_inst = inst.get("nombre", "Institución Educativa")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row, 1, nombre_inst.upper())
    cell.font = _INST_FONT
    cell.alignment = _INST_ALIGN
    row += 1

    sub_parts = []
    for key, label in [
        ("resolucion", "Resolución: "),
        ("DANE", "DANE: "),
        ("municipio", ""),
        ("direccion", ""),
        ("telefono", "Tel: "),
        ("rector", "Rector(a): "),
    ]:
        val = inst.get(key, "")
        if val:
            sub_parts.append(f"{label}{val}")
    if sub_parts:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row, 1, " — ".join(sub_parts))
        cell.font = _INST_SUB_FONT
        cell.alignment = _INST_SUB_ALIGN
        row += 1

    row += 1

    # ── 2. Título ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row, 1, "OBSERVADOR DEL ESTUDIANTE")
    cell.font = _TITULO_FONT
    cell.alignment = _TITULO_ALIGN
    row += 2

    # ── 3. Ficha del estudiante ─────────────────────────────────────────
    periodo_str = periodo or "Año completo"
    fecha_nac = est.get("fecha_nacimiento")
    fecha_nac_str = _fecha_corta(fecha_nac) if fecha_nac else "—"
    genero = est.get("genero") or "—"
    direccion_est = est.get("direccion") or "—"

    ficha = [
        ("ESTUDIANTE:", est.get("nombre", "—"), "DOCUMENTO:", est.get("documento", "—")),
        ("GRUPO:", est.get("grupo", "—"), "GRADO:", est.get("grado", "—")),
        ("FECHA NAC.:", fecha_nac_str, "GÉNERO:", genero),
        ("DIRECCIÓN:", direccion_est, "AÑO LECTIVO:", anio or "—"),
        ("PERIODO:", periodo_str, "", ""),
    ]

    def _write_ficha_row(r, label1, val1, label2, val2):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=6)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)

        c1 = ws.cell(r, 1, label1)
        c1.font = _LABEL_FONT
        c1.fill = _LABEL_FILL
        c1.border = _THIN_BORDER
        c2 = ws.cell(r, 3, _safe(val1))
        c2.font = _VALUE_FONT
        c2.alignment = _VALUE_ALIGN
        c2.border = _THIN_BORDER
        c3 = ws.cell(r, 6, label2)
        c3.font = _LABEL_FONT
        if label2:
            c3.fill = _LABEL_FILL
        c3.border = _THIN_BORDER
        c4 = ws.cell(r, 7, _safe(val2))
        c4.font = _VALUE_FONT
        c4.alignment = _VALUE_ALIGN
        c4.border = _THIN_BORDER

    for label1, val1, label2, val2 in ficha:
        _write_ficha_row(row, label1, val1, label2, val2)
        row += 1

    # ── 3b. Datos del acudiente principal ──────────────────────────────
    acud = est.get("acudiente") or {}
    if acud:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row, 1, "Datos del acudiente principal")
        cell.font = _SECCION_FONT
        cell.fill = _SECCION_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        ficha_acud = [
            ("ACUDIENTE:", acud.get("nombre", "—"), "PARENTESCO:", acud.get("parentesco_display", "—")),
            ("CELULAR:", acud.get("celular", "—"), "EMAIL:", acud.get("email", "—")),
            ("DIRECCIÓN:", acud.get("direccion", "—"), "DOCUMENTO:", acud.get("documento", "—")),
        ]
        for label1, val1, label2, val2 in ficha_acud:
            _write_ficha_row(row, label1, val1, label2, val2)
            row += 1

    row += 1

    # ── 4. Observaciones académicas ─────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row, 1, "I. Observaciones académicas")
    cell.font = _SECCION_FONT
    cell.fill = _SECCION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    obs = [e for e in entradas if e.get("tipo") == "observacion"]
    if not obs:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row, 1, "No se registran observaciones académicas en el periodo consultado.").font = Font(
            italic=True, size=9, color="718096"
        )
        row += 1
    else:
        obs_headers = ["N°", "Fecha y hora", "Periodo", "Asignatura", "Categoría", "Descripción", "Tipo", "", "Registrado por"]
        _write_row(ws, row, obs_headers, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_HEADER_ALIGN)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        header_row = row
        row += 1

        for i, o in enumerate(obs, 1):
            subtipo = o.get("subtipo", "")
            visib = "Pública" if subtipo == "publica" else "Privada"
            vals = [
                str(i),
                _fecha_hora(o.get("fecha")),
                _safe(o.get("periodo")),
                _safe(o.get("asignatura")),
                _safe(o.get("categoria")),
                _safe(o.get("descripcion")),
                visib,
                "",
                _safe(o.get("responsable")),
            ]
            fill = _zebra_fill(i)
            _write_row(ws, row, vals, font=_VALUE_FONT, alignment=_VALUE_ALIGN, fill=fill)
            ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
            row += 1

        _apply_borders(ws, header_row, row - 1, 9)

    row += 1

    # ── 5. Registros de comportamiento ──────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row, 1, "II. Registros de comportamiento")
    cell.font = _SECCION_FONT
    cell.fill = _SECCION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    regs = [e for e in entradas if e.get("tipo") == "registro"]
    if not regs:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row, 1, "No se registran anotaciones de comportamiento en el periodo consultado.").font = Font(
            italic=True, size=9, color="718096"
        )
        row += 1
    else:
        reg_headers = ["N°", "Fecha", "Periodo", "Tipo", "Situación", "Descripción", "Medida", "Seg.", "Registrado por"]
        _write_row(ws, row, reg_headers, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_HEADER_ALIGN)
        header_row = row
        row += 1

        for i, r in enumerate(regs, 1):
            subtipo = r.get("subtipo", "")
            n_seg = len(r.get("seguimiento_entries", []))
            vals = [
                str(i),
                _fecha_corta(r.get("fecha")),
                _safe(r.get("periodo")),
                _TIPO_LABEL.get(subtipo, subtipo),
                _safe(r.get("tipo_situacion")),
                _safe(r.get("descripcion")),
                _safe(r.get("medida")),
                str(n_seg) if n_seg else "—",
                _safe(r.get("responsable")),
            ]
            fill = _zebra_fill(i)
            _write_row(ws, row, vals, font=_VALUE_FONT, alignment=_VALUE_ALIGN, fill=fill)
            row += 1

        _apply_borders(ws, header_row, row - 1, 9)

    row += 1

    # ── 6. Detalle de seguimientos ──────────────────────────────────────
    regs_con_seg = [
        (i, e)
        for i, e in enumerate(
            (e for e in entradas if e.get("tipo") == "registro"), 1
        )
        if e.get("seguimiento_entries")
    ]
    if regs_con_seg:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row, 1, "Detalle de seguimientos")
        cell.font = _SECCION_FONT
        cell.fill = _SEG_HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        seg_headers = ["Reg. N°", "Fecha", "", "", "Anotación de seguimiento", "", "", "", "Responsable"]
        _write_row(ws, row, seg_headers, font=_HEADER_FONT, fill=_SEG_HEADER_FILL, alignment=_HEADER_ALIGN)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        header_row = row
        row += 1

        data_idx = 0
        for reg_num, entry in regs_con_seg:
            for seg in entry.get("seguimiento_entries", []):
                data_idx += 1
                vals = [
                    str(reg_num),
                    _fecha_hora(seg.get("fecha")),
                    "", "",
                    _safe(seg.get("texto")),
                    "", "", "",
                    _safe(seg.get("responsable")),
                ]
                fill = _zebra_fill(data_idx)
                _write_row(ws, row, vals, font=_VALUE_FONT, alignment=_VALUE_ALIGN, fill=fill)
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
                ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
                row += 1

        _apply_borders(ws, header_row, row - 1, 9)
        row += 1

    # ── 7. Resumen estadístico ──────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row, 1, "Resumen consolidado")
    cell.font = _SECCION_FONT
    cell.fill = _SECCION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    items = [
        ("Observaciones", resumen.get("num_observaciones", 0)),
        ("Fortalezas", resumen.get("fortalezas", 0)),
        ("Dificultades", resumen.get("dificultades", 0)),
        ("Compromisos", resumen.get("compromisos", 0)),
        ("Citaciones", resumen.get("citaciones", 0)),
        ("Descargos", resumen.get("descargos", 0)),
    ]

    # Labels
    col = 1
    for label, _ in items:
        color = _RESUMEN_COLORES.get(label, _AZUL)
        cell = ws.cell(row, col, label)
        cell.font = Font(bold=True, color=_BLANCO, size=9)
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
        col += 1
    row += 1

    # Values
    col = 1
    for _, val in items:
        cell = ws.cell(row, col, val)
        cell.font = Font(bold=True, size=11, color=_AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
        col += 1
    row += 2

    # ── 8. Notas de comportamiento por periodo ──────────────────────────
    notas = resumen.get("notas_por_periodo", {})
    if notas:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row, 1, "Valoración de comportamiento por periodo")
        cell.font = _SECCION_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=_AZUL_OSCURO)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        col = 1
        for per_nombre in notas:
            cell = ws.cell(row, col, per_nombre)
            cell.font = Font(bold=True, color=_BLANCO, size=9)
            cell.fill = PatternFill(fill_type="solid", fgColor=_AZUL_OSCURO)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER
            col += 1
        row += 1

        col = 1
        for val in notas.values():
            if val is not None:
                cell = ws.cell(row, col, float(val))
                cell.number_format = "0.0"
            else:
                cell = ws.cell(row, col, "—")
            cell.font = Font(bold=True, size=11, color=_AZUL)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER
            col += 1
        row += 2

    # ── 9. Pie con fecha de generación ──────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row, 1, f"Documento generado el {_date.today().strftime('%d/%m/%Y')} — Sistema ZECI Manager")
    cell.font = Font(size=8, italic=True, color="718096")
    cell.alignment = Alignment(horizontal="right")

    # ── Configurar impresión ────────────────────────────────────────────
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "portrait"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


__all__ = ["generar_observador_excel"]

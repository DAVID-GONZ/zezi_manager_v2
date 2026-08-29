"""
OpenpyxlExporter — implementación de IExporterService usando openpyxl.
"""

from __future__ import annotations

import io
from collections import Counter
from datetime import date as _date
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.domain.ports.service_ports import IExporterService

from .null_exporter import _csv_bytes

# Bins del histograma de notas — sincronizar con boletin_pdf.py
_NOTA_BINS = [(60, "<60"), (70, "60-69"), (80, "70-79"), (90, "80-89"), (float("inf"), "90-100")]


def _clasificar_notas(notas: list[float]) -> tuple[list[str], list[int]]:
    labels = [b[1] for b in _NOTA_BINS]
    counts = [0] * len(_NOTA_BINS)
    for n in notas:
        for i, (umbral, _) in enumerate(_NOTA_BINS):
            if n < umbral:
                counts[i] += 1
                break
    return labels, counts

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="2B6CB0")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT_ALIGN = Alignment(horizontal="right")
_FLOAT_FORMAT = "0.0"
_MAX_COL_WIDTH = 50

_AZUL = "2B6CB0"
_AZUL_CLARO = "DBEAFE"
_GRIS_CLARO = "F9FAFB"
_GRIS_BORDE = "CBD5E0"
_BLANCO = "FFFFFF"

_MEMBRETE_FONT = Font(bold=True, color=_AZUL, size=11)
_SUBTITLE_FONT = Font(bold=True, color="374151", size=9)
_NORMAL_FONT = Font(color="374151", size=9)
_BOLD_FONT = Font(bold=True, color="374151", size=9)
_SMALL_FONT = Font(color="718096", size=8)

_THIN_BORDER = Border(
    left=Side(style="thin", color=_GRIS_BORDE),
    right=Side(style="thin", color=_GRIS_BORDE),
    top=Side(style="thin", color=_GRIS_BORDE),
    bottom=Side(style="thin", color=_GRIS_BORDE),
)

_PIE_COLORS = ["2B6CB0", "38A169", "D69E2E", "E53E3E", "805AD5", "DD6B20", "319795", "B83280"]
_BAR_COLORS = ["38A169", "E53E3E", "D69E2E", "805AD5", "DD6B20", "3182CE"]


class OpenpyxlExporter(IExporterService):
    """
    Genera archivos Excel (.xlsx) con formato institucional.
    PDF no está implementado — usar un exportador especializado en HTML→PDF.
    CSV se sirve directamente sin openpyxl.
    """

    def exportar_excel(
        self,
        datos: list[dict],
        nombre_hoja: str = "Datos",
        ruta_destino: Path | None = None,
    ) -> bytes:
        """
        Crea un workbook Excel con:
          - Fila de headers en azul institucional (negrita, texto blanco)
          - Filas de datos con números alineados a la derecha
          - Ancho de columna auto-ajustado (máximo 50 chars)
        Retorna bytes si ruta_destino es None; escribe el archivo si se provee.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = nombre_hoja[:31]  # Excel limita el nombre de hoja a 31 chars

        if not datos:
            ws.cell(1, 1, "Sin datos")
        else:
            headers = list(datos[0].keys())

            # Fila de encabezados
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(1, col_idx, header)
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = _HEADER_ALIGN

            # Filas de datos
            for row_idx, fila in enumerate(datos, 2):
                for col_idx, key in enumerate(headers, 1):
                    valor = fila.get(key, "")
                    cell = ws.cell(row_idx, col_idx, valor)
                    if isinstance(valor, float):
                        cell.alignment = _RIGHT_ALIGN
                        cell.number_format = _FLOAT_FORMAT
                    elif isinstance(valor, int):
                        cell.alignment = _RIGHT_ALIGN

            # Auto-ancho de columnas (usa la línea más larga, no el total)
            for col_idx, header in enumerate(headers, 1):
                max_len = len(str(header))
                for fila in datos:
                    val = str(fila.get(header, ""))
                    longest_line = max((len(line) for line in val.split("\n")), default=0)
                    if longest_line > max_len:
                        max_len = longest_line
                ws.column_dimensions[get_column_letter(col_idx)].width = min(
                    max_len + 4, _MAX_COL_WIDTH
                )

        buffer = io.BytesIO()
        wb.save(buffer)
        contenido = buffer.getvalue()

        if ruta_destino is not None:
            Path(ruta_destino).write_bytes(contenido)
            return b""
        return contenido

    def exportar_pdf(
        self,
        html_content: str,
        ruta_destino: Path | None = None,
    ) -> bytes:
        raise NotImplementedError(
            "PDF no implementado en OpenpyxlExporter. "
            "Registra un exportador HTML→PDF (weasyprint, reportlab) para esta operación."
        )

    def exportar_csv(
        self,
        datos: list[dict],
        ruta_destino: Path | None = None,
        encoding: str = "utf-8-sig",
    ) -> bytes:
        contenido = _csv_bytes(datos, encoding)
        if ruta_destino is not None:
            Path(ruta_destino).write_bytes(contenido)
            return b""
        return contenido


# ── Reporte de convivencia por grupo (Excel enriquecido) ─────────────────────


def _auto_width(ws, col_idx: int, values: list, header: str, min_w: int = 8) -> None:
    max_len = max(len(str(header)), max((len(str(v)) for v in values), default=0))
    ws.column_dimensions[get_column_letter(col_idx)].width = max(
        min(max_len + 4, _MAX_COL_WIDTH), min_w
    )


def generar_reporte_convivencia_grupo_excel(
    filas: list[dict],
    titulo: str = "Reporte de convivencia",
    grupo: str = "",
    periodo: str = "",
    desglose_cols: list[str] | None = None,
) -> bytes:
    """Genera un Excel enriquecido del reporte de convivencia por grupo.

    Incluye:
    - Hoja "Reporte": membrete, tabla de datos con estilo, congelación de paneles.
    - Hoja "Estadísticos": resumen numérico, distribución por desempeño,
      totales de registros, pie chart, histograma de notas, bar chart de registros.
    """
    wb = openpyxl.Workbook()
    desglose_cols = desglose_cols or []

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 1: REPORTE
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Reporte"
    ws.sheet_properties.tabColor = _AZUL

    # ── Membrete ──
    ws.merge_cells("A1:F1")
    c_inst = ws.cell(1, 1, "INSTITUCIÓN EDUCATIVA ZECI")
    c_inst.font = _MEMBRETE_FONT
    c_inst.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:F2")
    c_titulo = ws.cell(2, 1, titulo)
    c_titulo.font = Font(bold=True, color=_AZUL, size=10)

    ws.merge_cells("A3:C3")
    ws.cell(3, 1, f"Curso: {grupo}   ·   {periodo}").font = _NORMAL_FONT

    ws.merge_cells("D3:F3")
    c_fecha = ws.cell(3, 4, f"Generado: {_date.today().strftime('%d/%m/%Y')}")
    c_fecha.font = _SMALL_FONT
    c_fecha.alignment = Alignment(horizontal="right")

    # Línea separadora
    for col in range(1, 12):
        ws.cell(4, col).border = Border(bottom=Side(style="medium", color=_AZUL))
    ws.row_dimensions[4].height = 6

    # ── Columnas del reporte ──
    base_cols: list[tuple[str, str]] = [
        ("estudiante", "Estudiante"),
        ("nota", "Nota comport."),
        ("nivel", "Desempeño"),
        ("fortalezas", "Fortalezas"),
        ("dificultades", "Dificultades"),
        ("compromisos", "Compromisos"),
        ("citaciones", "Citaciones"),
        ("descargos", "Descargos"),
    ]
    extra_cols = [(c, c) for c in desglose_cols]
    text_cols: list[tuple[str, str]] = [
        ("concepto", "Concepto de comportamiento"),
        ("observaciones", "Observaciones"),
    ]
    all_cols = base_cols + extra_cols + text_cols

    data_start_row = 6
    n_cols = len(all_cols)

    # Encabezados
    for col_idx, (_, header) in enumerate(all_cols, 1):
        cell = ws.cell(data_start_row, col_idx, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.row_dimensions[data_start_row].height = 24

    # Filas de datos
    alt_fill = PatternFill(fill_type="solid", fgColor=_GRIS_CLARO)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for row_offset, fila in enumerate(filas):
        row = data_start_row + 1 + row_offset
        is_alt = row_offset % 2 == 1
        for col_idx, (key, _) in enumerate(all_cols, 1):
            valor = fila.get(key, "")
            cell = ws.cell(row, col_idx, valor if valor != "" else None)
            cell.border = _THIN_BORDER
            if is_alt:
                cell.fill = alt_fill
            if key == "estudiante":
                cell.font = _NORMAL_FONT
                cell.alignment = left_align
            elif key in ("concepto", "observaciones"):
                cell.font = Font(color="374151", size=8)
                cell.alignment = wrap_align
            elif key == "nota":
                cell.font = _BOLD_FONT
                cell.alignment = center_align
                if isinstance(valor, float):
                    cell.number_format = "0.0"
            elif key == "nivel":
                cell.font = _NORMAL_FONT
                cell.alignment = center_align
            else:
                cell.font = _NORMAL_FONT
                cell.alignment = center_align

    last_data_row = data_start_row + len(filas)

    # Auto-ancho
    col_widths = {
        "estudiante": 28,
        "nota": 12,
        "nivel": 16,
        "concepto": 35,
        "observaciones": 40,
    }
    for col_idx, (key, header) in enumerate(all_cols, 1):
        w = col_widths.get(key, max(len(header) + 4, 14))
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Congelar paneles (encabezado visible al hacer scroll)
    ws.freeze_panes = f"A{data_start_row + 1}"

    # Filtro automático
    if filas:
        ws.auto_filter.ref = (
            f"A{data_start_row}:{get_column_letter(n_cols)}{last_data_row}"
        )

    # Configuración de impresión
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
        fitToPage=True
    )
    ws.print_title_rows = f"{data_start_row}:{data_start_row}"

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 2: ESTADÍSTICOS
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Estadísticos")
    ws2.sheet_properties.tabColor = "38A169"

    # ── Membrete ──
    ws2.merge_cells("A1:F1")
    ws2.cell(1, 1, f"ESTADÍSTICOS DE CONVIVENCIA — {grupo} · {periodo}").font = _MEMBRETE_FONT
    ws2.merge_cells("A2:F2")
    ws2.cell(2, 1, f"Generado: {_date.today().strftime('%d/%m/%Y')}").font = _SMALL_FONT
    for col in range(1, 8):
        ws2.cell(3, col).border = Border(bottom=Side(style="medium", color=_AZUL))
    ws2.row_dimensions[3].height = 6

    # Datos calculados
    total = len(filas)
    notas: list[float] = []
    for f in filas:
        n = f.get("nota")
        if n is not None and n != "":
            try:
                notas.append(float(n))
            except (ValueError, TypeError):
                pass
    promedio = round(sum(notas) / len(notas), 1) if notas else None
    nota_max = round(max(notas), 1) if notas else None
    nota_min = round(min(notas), 1) if notas else None
    sin_nota = total - len(notas)
    niveles = Counter(f.get("nivel", "") for f in filas if f.get("nivel"))
    total_fort = sum(f.get("fortalezas", 0) for f in filas)
    total_dif = sum(f.get("dificultades", 0) for f in filas)
    total_comp = sum(f.get("compromisos", 0) for f in filas)
    total_cit = sum(f.get("citaciones", 0) for f in filas)
    total_desc = sum(f.get("descargos", 0) for f in filas)
    total_obs = sum(f.get("num_obs", 0) for f in filas)

    # ── Tabla resumen general (A5:B11) ──
    ws2.cell(4, 1, "RESUMEN GENERAL").font = _SUBTITLE_FONT
    resumen_labels = [
        "Total estudiantes", "Con nota asignada", "Sin nota",
        "Promedio del grupo", "Nota más alta", "Nota más baja",
    ]
    resumen_values = [total, len(notas), sin_nota, promedio, nota_max, nota_min]
    for ci in (1, 2):
        cell = ws2.cell(5, ci, "Indicador" if ci == 1 else "Valor")
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    for i, (label, val) in enumerate(zip(resumen_labels, resumen_values)):
        r = 6 + i
        c_label = ws2.cell(r, 1, label)
        c_val = ws2.cell(r, 2, val if val is not None else "—")
        c_label.font = _NORMAL_FONT
        c_label.border = _THIN_BORDER
        c_val.border = _THIN_BORDER
        c_val.alignment = Alignment(horizontal="center")
        if label == "Promedio del grupo":
            c_val.font = _BOLD_FONT
        else:
            c_val.font = _NORMAL_FONT
        if isinstance(val, float):
            c_val.number_format = "0.0"
        if i % 2 == 1:
            c_label.fill = PatternFill(fill_type="solid", fgColor=_GRIS_CLARO)
            c_val.fill = PatternFill(fill_type="solid", fgColor=_GRIS_CLARO)

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14

    # ── Tabla distribución por desempeño (D5:F5+N) ──
    ws2.cell(4, 4, "DISTRIBUCIÓN POR DESEMPEÑO").font = _SUBTITLE_FONT
    dist_headers = ["Nivel de desempeño", "Cantidad", "%"]
    for ci, h in enumerate(dist_headers, 4):
        cell = ws2.cell(5, ci, h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    sorted_niveles = sorted(niveles.items(), key=lambda x: -x[1]) if niveles else []
    if not sorted_niveles:
        c = ws2.cell(6, 4, "Sin datos")
        c.font = _NORMAL_FONT
        c.border = _THIN_BORDER
        ws2.merge_cells("D6:F6")
        dist_last_row = 6
    else:
        for i, (nombre, cant) in enumerate(sorted_niveles):
            r = 6 + i
            pct = round(cant / total * 100) if total else 0
            ws2.cell(r, 4, nombre).font = _NORMAL_FONT
            ws2.cell(r, 5, cant).font = _NORMAL_FONT
            ws2.cell(r, 5).alignment = Alignment(horizontal="center")
            ws2.cell(r, 6, pct / 100).font = _NORMAL_FONT
            ws2.cell(r, 6).number_format = "0%"
            ws2.cell(r, 6).alignment = Alignment(horizontal="center")
            for ci in (4, 5, 6):
                ws2.cell(r, ci).border = _THIN_BORDER
                if i % 2 == 1:
                    ws2.cell(r, ci).fill = PatternFill(fill_type="solid", fgColor=_GRIS_CLARO)
        dist_last_row = 5 + len(sorted_niveles)

    ws2.column_dimensions["D"].width = 22
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 10

    # ── Pie chart: distribución por desempeño ──
    if sorted_niveles:
        pie = PieChart()
        pie.title = "Distribución por desempeño"
        pie.style = 10
        pie.width = 14
        pie.height = 10
        cats = Reference(ws2, min_col=4, min_row=6, max_row=dist_last_row)
        data = Reference(ws2, min_col=5, min_row=5, max_row=dist_last_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        pie.dataLabels.showCatName = False
        for i in range(len(sorted_niveles)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = _PIE_COLORS[i % len(_PIE_COLORS)]
            pie.series[0].data_points.append(pt)
        ws2.add_chart(pie, "H4")

    # ── Tabla totales de registros ──
    reg_row_start = max(dist_last_row, 11) + 2
    ws2.cell(reg_row_start, 1, "TOTALES DE REGISTROS").font = _SUBTITLE_FONT
    reg_headers = ["Fortalezas", "Dificultades", "Compromisos", "Citaciones", "Descargos", "Observaciones"]
    reg_values = [total_fort, total_dif, total_comp, total_cit, total_desc, total_obs]

    for ci, h in enumerate(reg_headers, 1):
        cell = ws2.cell(reg_row_start + 1, ci, h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    for ci, v in enumerate(reg_values, 1):
        cell = ws2.cell(reg_row_start + 2, ci, v)
        cell.font = _BOLD_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    # ── Bar chart: registros por tipo ──
    bar = BarChart()
    bar.type = "col"
    bar.title = "Registros por tipo"
    bar.style = 10
    bar.width = 18
    bar.height = 10
    bar.y_axis.title = "Cantidad"
    cats_bar = Reference(ws2, min_col=1, min_row=reg_row_start + 1, max_col=6)
    data_bar = Reference(ws2, min_col=1, min_row=reg_row_start + 2, max_col=6)
    bar.add_data(data_bar, from_rows=True, titles_from_data=False)
    bar.set_categories(cats_bar)
    bar.shape = 4
    bar.legend = None
    if bar.series:
        s = bar.series[0]
        for i in range(min(6, len(reg_values))):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = _BAR_COLORS[i % len(_BAR_COLORS)]
            s.data_points.append(pt)
        s.graphicalProperties.line.noFill = True
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    ws2.add_chart(bar, f"A{reg_row_start + 4}")

    # ── Histograma de notas ──
    hist_row = reg_row_start + 4
    ws2.cell(hist_row, 5, "DISTRIBUCIÓN DE NOTAS").font = _SUBTITLE_FONT
    bin_labels, counts = _clasificar_notas(notas)

    for ci, (label, cnt) in enumerate(zip(bin_labels, counts), 5):
        ws2.cell(hist_row + 1, ci, label).font = _SMALL_FONT
        ws2.cell(hist_row + 1, ci).alignment = Alignment(horizontal="center")
        ws2.cell(hist_row + 1, ci).border = _THIN_BORDER
        ws2.cell(hist_row + 1, ci).fill = PatternFill(fill_type="solid", fgColor=_AZUL)
        ws2.cell(hist_row + 1, ci).font = Font(bold=True, color=_BLANCO, size=8)
        ws2.cell(hist_row + 2, ci, cnt).font = _BOLD_FONT
        ws2.cell(hist_row + 2, ci).alignment = Alignment(horizontal="center")
        ws2.cell(hist_row + 2, ci).border = _THIN_BORDER

    if notas:
        hist_chart = BarChart()
        hist_chart.type = "col"
        hist_chart.title = "Distribución de notas"
        hist_chart.style = 10
        hist_chart.width = 14
        hist_chart.height = 10
        hist_chart.y_axis.title = "Estudiantes"
        cats_hist = Reference(ws2, min_col=5, min_row=hist_row + 1, max_col=9)
        data_hist = Reference(ws2, min_col=5, min_row=hist_row + 2, max_col=9)
        hist_chart.add_data(data_hist, from_rows=True, titles_from_data=False)
        hist_chart.set_categories(cats_hist)
        hist_chart.legend = None
        if hist_chart.series:
            s = hist_chart.series[0]
            s.graphicalProperties.solidFill = "3182CE"
            s.graphicalProperties.line.solidFill = "2B6CB0"
        hist_chart.dataLabels = DataLabelList()
        hist_chart.dataLabels.showVal = True
        ws2.add_chart(hist_chart, f"H{reg_row_start + 4}")

    # Configuración de impresión hoja 2
    ws2.page_setup.orientation = "landscape"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["OpenpyxlExporter", "generar_reporte_convivencia_grupo_excel"]

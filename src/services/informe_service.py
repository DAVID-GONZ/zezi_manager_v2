"""
InformeService
===============
Orquesta la generación y exportación de informes académicos.

Coordina el repositorio de estadísticos (obtención de datos) con el
IExporterService (conversión al formato de salida). No contiene lógica
de presentación ni accede directamente a la BD.
"""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass, field
from typing import TYPE_CHECKING

from src.domain.models.dtos import (
    FormatoInforme,
    InformeAsistenciaDTO,
    InformeNotasDTO,
)
from src.domain.ports.estadisticos_repo import IEstadisticosRepository
from src.domain.ports.service_ports import IExporterService

if TYPE_CHECKING:
    from src.services.convivencia_service import ConvivenciaService

# ── Sanitización de datos para exportación ───────────────────────────────────

# Campos que nunca deben aparecer en un documento exportado
_CAMPOS_EXCLUIR: frozenset[str] = frozenset(
    {
        "estudiante_id",
        # Los flags *_perdio del consolidado anual son internos
    }
)

# Renombre de claves de sistema a nombres semánticos en español
_CAMPO_RENOMBRAR: dict[str, str] = {
    "nombre_completo": "Estudiante",
    "documento": "Documento",
    "nombre_asignatura": "Asignatura",
    "promedio_periodo": "Promedio",
    "promedio": "Promedio",
    "posicion": "#",
    "presentes": "Presentes",
    "faltas_injustificadas": "F. Injustificadas",
    "faltas_justificadas": "F. Justificadas",
    "retrasos": "Retrasos",
    "excusas": "Excusas",
    "porcentaje": "% Asistencia",
    "estado_promocion": "Estado",
    "definitiva": "Definitiva",
    "area_nombre": "Área",
    "total_asignaturas": "Asignaturas",
}


@dataclass
class BoletinesGrupoDTO:
    """Resultado de la generación masiva de boletines de un grupo."""

    contenido: bytes | None = None  # PDF/Excel fusionado (None si nada)
    errores: list[str] = field(default_factory=list)  # nombres con fallo


def sanitizar_datos_exportacion(datos: list[dict]) -> list[dict]:
    """
    Prepara datos brutos del repositorio para exportación (Excel / PDF):

    - Elimina ``estudiante_id`` y cualquier columna que termine en ``_perdio``
      (flags internos del consolidado anual).
    - Renombra claves de sistema a nombres semánticos en español usando
      ``_CAMPO_RENOMBRAR``.  Las claves sin mapeo se conservan tal cual
      (típicamente son nombres de asignaturas ya legibles).

    Se aplica solo al exportar; las vistas ag-Grid del navegador usan los
    nombres originales (``field`` en columnDefs).
    """
    if not datos:
        return datos
    resultado: list[dict] = []
    for fila in datos:
        nueva: dict = {}
        for clave, valor in fila.items():
            if clave in _CAMPOS_EXCLUIR or clave.endswith("_perdio"):
                continue
            nueva[_CAMPO_RENOMBRAR.get(clave, clave)] = valor
        resultado.append(nueva)
    return resultado


# ─────────────────────────────────────────────────────────────────────────────


class InformeService:
    """
    Orquesta la generación de informes académicos en diferentes formatos.
    No contiene SQL. No contiene lógica de presentación.
    """

    def __init__(
        self,
        estadisticos_repo: IEstadisticosRepository,
        exporter: IExporterService | None = None,
        estudiante_repo=None,
        convivencia_svc_provider: Callable[[], ConvivenciaService] | None = None,
    ) -> None:
        """Inyecta el repo de estadísticos y, opcionalmente, el exportador,
        el repo de estudiantes y un proveedor lazy de ``ConvivenciaService``.

        ``convivencia_svc_provider`` es un callable que retorna
        ``ConvivenciaService``; se usa lazy para evitar ciclos de wiring.
        Si es None, los métodos de convivencia retornan dicts vacíos.
        """
        self._estadisticos_repo = estadisticos_repo
        self._exporter = exporter
        self._estudiante_repo = estudiante_repo
        self._convivencia_svc_provider = convivencia_svc_provider

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_exporter_o_lanzar(self) -> IExporterService:
        if self._exporter is None:
            raise ValueError(
                "No hay un exportador configurado. "
                "Proporcione una implementación de IExporterService."
            )
        return self._exporter

    # ------------------------------------------------------------------
    # Datos de convivencia para boletín (convivencia_32 — thin wrappers)
    # ------------------------------------------------------------------

    def _conv_svc(self):
        """Retorna el ConvivenciaService lazy, o None si no hay provider."""
        if self._convivencia_svc_provider is None:
            return None
        return self._convivencia_svc_provider()

    def convivencia_boletin(
        self,
        estudiante_id: int,
        periodo_id: int,
    ) -> dict:
        """Thin wrapper — delega en ConvivenciaService.paquete_boletin_periodo.

        Claves del dict retornado:
          nota:                        float | None
          nota_observacion:            str   | None
          observaciones:               list[str]   (textos planos — compat retro PDF)
          observaciones_por_categoria: list[dict]  (formato rico)
          registros:                   list[dict]  (política convivencia_29)

        Si no hay provider → retorna dict con None/[].
        """
        svc = self._conv_svc()
        if svc is None:
            return {
                "nota": None,
                "nota_observacion": None,
                "observaciones": [],
                "registros": [],
            }
        return svc.paquete_boletin_periodo(estudiante_id, periodo_id)

    def convivencia_boletin_anual(
        self,
        estudiante_id: int,
        anio_id: int,
    ) -> dict:
        """Thin wrapper — delega en ConvivenciaService.paquete_boletin_anual.

        Si no hay provider → retorna estructura vacía sin lanzar excepción.
        """
        svc = self._conv_svc()
        if svc is None:
            return {
                "periodos": [],
                "notas_por_periodo": {},
                "definitiva": None,
                "concepto": None,
                "observaciones_por_categoria": [],
                "registros": [],
            }
        return svc.paquete_boletin_anual(estudiante_id, anio_id)

    # ------------------------------------------------------------------
    # Informe de notas
    # ------------------------------------------------------------------

    def datos_informe_notas(
        self,
        dto: InformeNotasDTO,
    ) -> list[dict]:
        """
        Obtiene los datos del informe de notas para un grupo y periodo.

        Retorna una lista de dicts donde cada fila es un estudiante
        con sus notas definitivas por asignatura.
        """
        return self._estadisticos_repo.consolidado_notas_grupo(dto.grupo_id, dto.periodo_id)

    def generar_notas(
        self,
        dto: InformeNotasDTO,
        contexto: dict | None = None,
    ) -> bytes:
        """
        Genera el informe de notas en el formato especificado (Excel o PDF).

        Obtiene los datos del consolidado y los exporta al formato indicado.
        Retorna el contenido como bytes para descarga directa.

        ``contexto`` puede traer grupo_nombre / periodo_nombre / asignatura_nombre
        para el membrete del PDF.

        Lanza:
            ValueError: Si no hay exportador configurado.
        """
        exporter = self._get_exporter_o_lanzar()
        datos = sanitizar_datos_exportacion(self.datos_informe_notas(dto))

        if dto.formato == FormatoInforme.EXCEL:
            return exporter.exportar_excel(
                datos,
                nombre_hoja=f"Notas Periodo {dto.periodo_id}",
            )
        else:
            html = self._datos_a_html(
                datos,
                titulo=f"Informe de Notas — Periodo {dto.periodo_id}",
            )
            return exporter.exportar_pdf(self._inyectar_meta_html(html, contexto))

    # ------------------------------------------------------------------
    # Informe de asistencia
    # ------------------------------------------------------------------

    def datos_informe_asistencia(
        self,
        dto: InformeAsistenciaDTO,
    ) -> list[dict]:
        """
        Obtiene los datos del informe de asistencia para un grupo y periodo.

        Retorna una lista de dicts donde cada fila es un registro de
        asistencia por estudiante y asignatura.
        """
        return self._estadisticos_repo.consolidado_asistencia_grupo(dto.grupo_id, dto.periodo_id)

    def generar_asistencia(
        self,
        dto: InformeAsistenciaDTO,
        contexto: dict | None = None,
    ) -> bytes:
        """
        Genera el informe de asistencia en el formato especificado.

        Retorna el contenido como bytes para descarga directa.

        ``contexto`` puede traer grupo_nombre / periodo_nombre para el
        membrete del PDF.

        Lanza:
            ValueError: Si no hay exportador configurado.
        """
        exporter = self._get_exporter_o_lanzar()
        datos = sanitizar_datos_exportacion(self.datos_informe_asistencia(dto))

        if dto.formato == FormatoInforme.EXCEL:
            return exporter.exportar_excel(
                datos,
                nombre_hoja=f"Asistencia Periodo {dto.periodo_id}",
            )
        else:
            html = self._datos_a_html(
                datos,
                titulo=f"Informe de Asistencia — Periodo {dto.periodo_id}",
            )
            return exporter.exportar_pdf(self._inyectar_meta_html(html, contexto))

    # ------------------------------------------------------------------
    # Informe anual consolidado
    # ------------------------------------------------------------------

    def datos_consolidado_anual(
        self,
        grupo_id: int,
        anio_id: int,
    ) -> list[dict]:
        """
        Obtiene el consolidado anual: notas + estado de promoción.

        Usado para generar el acta final de calificaciones del año.
        """
        return self._estadisticos_repo.consolidado_anual_grupo(grupo_id, anio_id)

    def generar_consolidado_anual(
        self,
        grupo_id: int,
        anio_id: int,
        formato: FormatoInforme = FormatoInforme.EXCEL,
    ) -> bytes:
        """
        Genera el consolidado anual en el formato especificado.

        Retorna el contenido como bytes para descarga directa.

        Lanza:
            ValueError: Si no hay exportador configurado.
        """
        exporter = self._get_exporter_o_lanzar()
        datos = sanitizar_datos_exportacion(self.datos_consolidado_anual(grupo_id, anio_id))

        if formato == FormatoInforme.EXCEL:
            return exporter.exportar_excel(
                datos,
                nombre_hoja=f"Consolidado Anual {anio_id}",
            )
        else:
            html = self._datos_a_html(
                datos,
                titulo=f"Consolidado Anual — Año {anio_id}",
            )
            return exporter.exportar_pdf(html)

    # ------------------------------------------------------------------
    # Exportación directa de datos
    # ------------------------------------------------------------------

    def exportar_csv(
        self,
        datos: list[dict],
        encoding: str = "utf-8-sig",
    ) -> bytes:
        """
        Exporta una lista de dicts como CSV.

        Útil para integraciones con otros sistemas.
        Lanza si no hay exportador configurado.
        """
        exporter = self._get_exporter_o_lanzar()
        return exporter.exportar_csv(datos, encoding=encoding)

    # ------------------------------------------------------------------
    # Boletín por periodo (individual por estudiante)
    # ------------------------------------------------------------------

    def generar_boletin_periodo(
        self,
        estudiante_id: int,
        grupo_id: int,
        periodo_id: int,
        formato: str = "pdf",
        grupo_nombre: str = "",
        periodo_nombre: str = "",
    ) -> bytes:
        """
        Genera el boletín de un estudiante para un periodo específico.

        PDF: genera un documento formal con membrete, tabla Área > Asignatura,
             asistencia por tipo, observaciones y firmas.
        Excel: tabla plana con nota y asistencia por asignatura.

        Lanza:
            ValueError: Si no hay exportador configurado.
        """
        fmt = FormatoInforme(formato)

        if fmt == FormatoInforme.PDF:
            import importlib

            _boletin_mod = importlib.import_module("src.infrastructure.exporters.boletin_pdf")
            datos = self._estadisticos_repo.boletin_datos_acumulado(
                estudiante_id, grupo_id, periodo_id
            )
            datos["convivencia"] = self.convivencia_boletin(estudiante_id, periodo_id)
            return _boletin_mod.generar_boletin_acumulado_pdf(datos)

        # Excel: tabla plana acumulada con columna por cada periodo anterior + actual
        # (convivencia_31) Las columnas de convivencia se llevan a la hoja "Convivencia".
        exporter = self._get_exporter_o_lanzar()
        datos_raw = self._estadisticos_repo.boletin_datos_acumulado(
            estudiante_id, grupo_id, periodo_id
        )
        periodos = datos_raw.get("periodos", [])
        label_def = "Definitiva" if datos_raw.get("es_ultimo_periodo") else "Promedio"
        filas: list[dict] = []
        for area in datos_raw.get("areas", []):
            for asig in area.get("asignaturas", []):
                fila: dict = {
                    "Área": area["area_nombre"],
                    "Asignatura": asig["nombre"],
                }
                for per in periodos:
                    fila[per["nombre"]] = asig.get("notas_periodo", {}).get(per["id"])
                fila[label_def] = asig.get("definitiva")
                fila["Presentes"] = asig.get("presentes", 0)
                fila["F. Inj."] = asig.get("faltas_injustificadas", 0)
                fila["F. Just."] = asig.get("faltas_justificadas", 0)
                fila["Retrasos"] = asig.get("retrasos", 0)
                fila["Excusas"] = asig.get("excusas", 0)
                filas.append(fila)
        main_bytes = exporter.exportar_excel(filas, nombre_hoja="Boletín Periodo")
        if self._conv_svc() is None:
            return main_bytes  # R5: sin provider no se añade hoja de convivencia
        conv_filas = self._hoja_convivencia_periodo(estudiante_id, periodo_id)
        conv_bytes = self._serializar_hoja_matriz(conv_filas, "Convivencia")
        return merge_excels([("Boletín Periodo", main_bytes), ("Convivencia", conv_bytes)])

    # ------------------------------------------------------------------
    # Boletín anual (individual por estudiante)
    # ------------------------------------------------------------------

    def generar_boletin_anual(
        self,
        estudiante_id: int,
        grupo_id: int,
        anio_id: int,
        formato: str = "pdf",
        grupo_nombre: str = "",
    ) -> bytes:
        """
        Genera el boletín anual de un estudiante.

        PDF: documento formal con tabla Área > Asignatura, columnas por periodo
             configuradas dinámicamente, definitiva, asistencia anual, firmas.
        Excel: un libro con una hoja por área, filas por asignatura.

        Lanza:
            ValueError: Si no hay exportador configurado.
        """
        fmt = FormatoInforme(formato)

        if fmt == FormatoInforme.PDF:
            import importlib

            _boletin_mod = importlib.import_module("src.infrastructure.exporters.boletin_pdf")
            datos = self._estadisticos_repo.boletin_datos_anual(estudiante_id, grupo_id, anio_id)
            datos["convivencia_anual"] = self.convivencia_boletin_anual(estudiante_id, anio_id)
            return _boletin_mod.generar_boletin_anual_pdf(datos)

        # Excel: tabla plana con columna por periodo
        # (convivencia_31) Las columnas de convivencia se llevan a la hoja "Convivencia".
        exporter = self._get_exporter_o_lanzar()
        datos_raw = self._estadisticos_repo.boletin_datos_anual(estudiante_id, grupo_id, anio_id)
        periodos = datos_raw.get("periodos", [])

        filas: list[dict] = []
        for area in datos_raw.get("areas", []):
            for asig in area.get("asignaturas", []):
                fila: dict = {
                    "Área": area["area_nombre"],
                    "Asignatura": asig["nombre"],
                }
                for per in periodos:
                    nota = asig.get("notas_periodo", {}).get(per["id"])
                    fila[per["nombre"]] = nota
                fila["Definitiva"] = asig.get("definitiva")
                fila["Presentes"] = asig.get("presentes", 0)
                fila["F. Inj."] = asig.get("faltas_injustificadas", 0)
                fila["F. Just."] = asig.get("faltas_justificadas", 0)
                fila["Retrasos"] = asig.get("retrasos", 0)
                fila["Excusas"] = asig.get("excusas", 0)
                filas.append(fila)
        main_bytes = exporter.exportar_excel(filas, nombre_hoja="Boletín Anual")
        if self._conv_svc() is None:
            return main_bytes  # R5: sin provider no se añade hoja de convivencia
        conv_filas = self._hoja_convivencia_anual(estudiante_id, anio_id)
        conv_bytes = self._serializar_hoja_matriz(conv_filas, "Convivencia")
        return merge_excels([("Boletín Anual", main_bytes), ("Convivencia", conv_bytes)])

    # ------------------------------------------------------------------
    # Boletines masivos por grupo (genera + fusiona, sin merge en la vista)
    # ------------------------------------------------------------------

    def generar_boletines_grupo(
        self,
        grupo_id: int,
        periodo_id: int | None = None,
        anio_id: int | None = None,
        formato: str = "pdf",
        grupo_nombre: str = "",
        periodo_nombre: str = "",
    ) -> BoletinesGrupoDTO:
        """Genera el boletín de cada estudiante del grupo y los fusiona en un
        único documento (PDF combinado o Excel con una hoja por estudiante).

        Modo periodo: pasar `periodo_id`. Modo anual: pasar `anio_id`.
        Devuelve el documento fusionado y la lista de estudiantes con error.
        """
        if periodo_id is None and anio_id is None:
            raise ValueError("Debe indicar periodo_id (boletín de periodo) o anio_id (anual).")
        if self._estudiante_repo is None:
            raise ValueError("InformeService no tiene estudiante_repo configurado.")

        from src.services.contexto_tenant import institucion_actual
        fmt = FormatoInforme(formato)
        estudiantes = self._estudiante_repo.listar_por_grupo(grupo_id, institucion_actual() or "*")

        pdfs: list[bytes] = []
        hojas: list[tuple[str, bytes]] = []
        errores: list[str] = []

        for est in estudiantes:
            nombre = f"{est.nombre} {est.apellido}"
            try:
                if anio_id is not None:
                    contenido = self.generar_boletin_anual(
                        est.id,
                        grupo_id,
                        anio_id,
                        formato,
                        grupo_nombre=grupo_nombre,
                    )
                else:
                    contenido = self.generar_boletin_periodo(
                        est.id,
                        grupo_id,
                        periodo_id,
                        formato,
                        grupo_nombre=grupo_nombre,
                        periodo_nombre=periodo_nombre,
                    )
            except Exception:
                errores.append(nombre)
                continue
            if fmt == FormatoInforme.PDF:
                pdfs.append(contenido)
            else:
                hojas.append((f"{est.apellido} {est.nombre}"[:31], contenido))

        if fmt == FormatoInforme.PDF:
            contenido_final = merge_pdfs(pdfs) if pdfs else None
        else:
            # (convivencia_31) Añadir hoja resumen "Convivencia — todos" si hay provider.
            if hojas and self._conv_svc() is not None:
                resumen_filas = self._hoja_convivencia_resumen_grupo(grupo_id, periodo_id, anio_id)
                resumen_bytes = self._serializar_hoja_matriz(resumen_filas, "Convivencia — todos")
                hojas.append(("Convivencia — todos", resumen_bytes))
            contenido_final = merge_excels(hojas) if hojas else None

        return BoletinesGrupoDTO(contenido=contenido_final, errores=errores)

    # ------------------------------------------------------------------
    # Helpers privados — hojas de convivencia (convivencia_31)
    # ------------------------------------------------------------------

    @staticmethod
    def _serializar_hoja_matriz(filas: list[list], nombre: str) -> bytes:
        """Serializa una matriz irregular (list[list]) como xlsx de una sola hoja.

        No pasa por el puerto IExporterService; usa openpyxl directamente para
        soportar estructuras heterogéneas (filas con distinto número de celdas).
        """
        import io as _io

        import openpyxl as _xl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = _xl.Workbook()
        ws = wb.active
        ws.title = nombre[:31]

        hdr_font = Font(bold=True, color="FFFFFF", size=9)
        hdr_fill = PatternFill(fill_type="solid", fgColor="2B6CB0")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="CBD5E0"),
            right=Side(style="thin", color="CBD5E0"),
            top=Side(style="thin", color="CBD5E0"),
            bottom=Side(style="thin", color="CBD5E0"),
        )

        max_cols = 0
        for fila in filas:
            ws.append(fila if fila else [""])
            if fila:
                max_cols = max(max_cols, len(fila))

        if filas:
            for cell in ws[1]:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_cols or 1):
            for cell in row:
                cell.border = thin_border

        for col_idx in range(1, (max_cols or 1) + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    longest = max((len(line) for line in val.split("\n")), default=0)
                    if longest > max_len:
                        max_len = longest
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _hoja_convivencia_periodo(self, estudiante_id: int, periodo_id: int) -> list[list]:
        """Matriz para la hoja 'Convivencia' del boletín de periodo individual."""
        conv = self.convivencia_boletin(estudiante_id, periodo_id)
        filas: list[list] = [["CONVIVENCIA"], []]
        filas.append(["Nota de comportamiento", conv["nota"] if conv["nota"] is not None else "—"])
        filas.append(["Concepto", conv.get("nota_observacion") or ""])
        filas.append([])
        # Observaciones agrupadas por categoría (disponibles desde convivencia_32;
        # antes de ese paso la clave puede no estar presente → bloque vacío).
        for grupo in conv.get("observaciones_por_categoria", []):
            if not grupo["items"]:
                continue
            filas.append([grupo["categoria"]])
            filas.append(["Fecha", "Autor", "Texto"])
            for it in grupo["items"]:
                filas.append([it.get("fecha", ""), it.get("autor", ""), it.get("texto", "")])
            filas.append([])
        # Registros de comportamiento (convivencia_29)
        if conv.get("registros"):
            filas.append(["Eventos"])
            filas.append(["Fecha", "Tipo", "Descripción"])
            for r in conv["registros"]:
                filas.append([r["fecha"], r["tipo"], r["descripcion"]])
        return filas

    def _hoja_convivencia_anual(self, estudiante_id: int, anio_id: int) -> list[list]:
        """Matriz para la hoja 'Convivencia' del boletín anual individual."""
        conv = self.convivencia_boletin_anual(estudiante_id, anio_id)
        filas: list[list] = [["CONVIVENCIA ANUAL"], []]
        # Tabla resumen por periodo
        filas.append(["Periodo", "Nota", "Concepto"])
        notas_por_periodo = conv.get("notas_por_periodo", {})
        for per in conv.get("periodos", []):
            nota = notas_por_periodo.get(per["id"])
            filas.append([per["nombre"], nota if nota is not None else "—", ""])
        filas.append(
            [
                "Definitiva",
                conv["definitiva"] if conv["definitiva"] is not None else "—",
                conv.get("concepto") or "",
            ]
        )
        filas.append([])
        # Observaciones agrupadas por categoría
        for grupo in conv.get("observaciones_por_categoria", []):
            if not grupo["items"]:
                continue
            filas.append([grupo["categoria"]])
            filas.append(["Periodo", "Autor", "Texto"])
            for it in grupo["items"]:
                filas.append(
                    [
                        it.get("periodo", ""),
                        it.get("autor", ""),
                        it.get("texto", ""),
                    ]
                )
            filas.append([])
        # Registros del año (convivencia_29)
        if conv.get("registros"):
            filas.append(["Eventos"])
            filas.append(["Fecha", "Tipo", "Descripción"])
            for r in conv["registros"]:
                filas.append([r["fecha"], r["tipo"], r["descripcion"]])
        return filas

    def _hoja_convivencia_resumen_grupo(
        self,
        grupo_id: int,
        periodo_id: int | None,
        anio_id: int | None,
    ) -> list[list]:
        """Matriz para la hoja 'Convivencia — todos' del libro masivo.

        Modo periodo (periodo_id not None): encabezado Estudiante|Nota|Concepto|
        #Obs.|#Eventos + pies de obs y registros.
        Modo anual (anio_id not None): encabezado Estudiante|P1|P2|…|Definitiva|
        Concepto final|#Obs.|#Eventos.

        Si no hay convivencia_repo o estudiante_repo → fila vacía con aviso.
        """
        _svc = self._conv_svc()
        if _svc is None or self._estudiante_repo is None:
            return [["Sin datos de convivencia disponibles"]]
        # Acceso al repo a través del servicio (convivencia_32: ya no hay _convivencia_repo).
        _repo = _svc._repo

        from src.domain.models.convivencia import (
            TIPO_REGISTRO_DISPLAY,
            FiltroConvivenciaDTO,
        )

        from src.services.contexto_tenant import institucion_actual
        estudiantes = sorted(
            self._estudiante_repo.listar_por_grupo(grupo_id, institucion_actual() or "*"),
            key=lambda e: f"{getattr(e, 'apellido', '')} {getattr(e, 'nombre', '')}".lower(),
        )
        est_nombre = {
            e.id: f"{getattr(e, 'apellido', '')} {getattr(e, 'nombre', '')}".strip()
            for e in estudiantes
        }

        filas: list[list] = []

        if periodo_id is not None:
            # ── Modo periodo ───────────────────────────────────────────
            filas.append([f"CONVIVENCIA — Grupo {grupo_id} — Periodo {periodo_id}"])
            filas.append([])
            filas.append(["Estudiante", "Nota", "Concepto", "# Obs. públicas", "# Eventos"])

            # Batch: observaciones públicas del grupo en el periodo
            try:
                obs_all = _repo.listar_observaciones_por_grupo(
                    grupo_id, periodo_id, solo_publicas=True
                )
            except Exception:
                obs_all = []
            obs_count: dict[int, int] = {}
            for o in obs_all:
                obs_count[o.estudiante_id] = obs_count.get(o.estudiante_id, 0) + 1

            # Batch: registros del grupo en el periodo
            try:
                from src.services.contexto_tenant import institucion_actual

                regs_all = _repo.listar_registros(
                    FiltroConvivenciaDTO(grupo_id=grupo_id, periodo_id=periodo_id),
                    institucion_id=institucion_actual() or "*",
                )
            except Exception:
                regs_all = []
            reg_count: dict[int, int] = {}
            for r in regs_all:
                reg_count[r.estudiante_id] = reg_count.get(r.estudiante_id, 0) + 1

            for est in estudiantes:
                nota_obj = _repo.get_nota(est.id, periodo_id)
                nota_val = nota_obj.valor if nota_obj is not None else None
                concepto = (nota_obj.observacion if nota_obj is not None else None) or ""
                filas.append(
                    [
                        est_nombre.get(est.id, ""),
                        nota_val if nota_val is not None else "—",
                        concepto,
                        obs_count.get(est.id, 0),
                        reg_count.get(est.id, 0),
                    ]
                )

            # Pie — Observaciones públicas (todas)
            if obs_all:
                filas.append([])
                filas.append(["Observaciones públicas (todas)"])
                filas.append(["Estudiante", "Categoría", "Fecha", "Autor", "Texto"])
                for o in sorted(
                    obs_all,
                    key=lambda x: (est_nombre.get(x.estudiante_id, ""), str(x.categoria_id or "")),
                ):
                    fecha_str = (
                        str(o.fecha_registro.date())
                        if hasattr(o.fecha_registro, "date")
                        else str(o.fecha_registro)
                    )
                    filas.append(
                        [
                            est_nombre.get(o.estudiante_id, ""),
                            str(o.categoria_id or ""),
                            fecha_str,
                            "",  # autor: ObservacionPeriodo solo tiene usuario_id (migrar en 32)
                            o.texto,
                        ]
                    )

            # Pie — Eventos de convivencia
            if regs_all:
                filas.append([])
                filas.append(["Eventos de convivencia"])
                filas.append(["Estudiante", "Fecha", "Tipo", "Descripción"])
                for r in sorted(
                    regs_all,
                    key=lambda x: (est_nombre.get(x.estudiante_id, ""), str(x.fecha)),
                ):
                    filas.append(
                        [
                            est_nombre.get(r.estudiante_id, ""),
                            str(r.fecha),
                            TIPO_REGISTRO_DISPLAY.get(r.tipo.value, r.tipo.value),
                            r.descripcion,
                        ]
                    )

        else:
            # ── Modo anual ─────────────────────────────────────────────
            periodos = []
            if _svc._periodo_svc_provider is not None:
                try:
                    periodos = _svc._periodo_svc_provider().listar_por_anio(anio_id)
                except Exception:
                    periodos = []

            filas.append([f"CONVIVENCIA — Grupo {grupo_id} — Año {anio_id}"])
            filas.append([])
            per_headers = [p.nombre for p in periodos]
            filas.append(
                ["Estudiante", *per_headers, "Definitiva", "Concepto final", "# Obs. públicas", "# Eventos"]
            )

            for est in estudiantes:
                notas_est = {n.periodo_id: n for n in _repo.listar_notas_por_estudiante(est.id)}
                notas_vals = [
                    (notas_est[p.id].valor if p.id in notas_est else None) for p in periodos
                ]
                notas_presentes = [v for v in notas_vals if v is not None]
                definitiva = (
                    round(sum(notas_presentes) / len(notas_presentes), 2)
                    if notas_presentes
                    else None
                )
                # Concepto del último periodo con nota
                concepto_final = ""
                for p in reversed(periodos):
                    if p.id in notas_est and notas_est[p.id].observacion:
                        concepto_final = notas_est[p.id].observacion
                        break

                # Contar obs y eventos del año
                obs_cnt = sum(
                    len(_repo.listar_observaciones_por_estudiante(est.id, p.id, solo_publicas=True))
                    for p in periodos
                )
                from src.services.contexto_tenant import institucion_actual

                _inst_scope = institucion_actual() or "*"
                reg_cnt = sum(
                    len(
                        _repo.listar_registros(
                            FiltroConvivenciaDTO(estudiante_id=est.id, periodo_id=p.id),
                            institucion_id=_inst_scope,
                        )
                    )
                    for p in periodos
                )

                fila_vals = [(v if v is not None else "—") for v in notas_vals]
                filas.append(
                    [est_nombre.get(est.id, ""), *fila_vals, definitiva if definitiva is not None else "—", concepto_final, obs_cnt, reg_cnt]
                )

        return filas

    # ------------------------------------------------------------------
    # Exportación de estadísticos (encapsula el pipeline de la vista)
    # ------------------------------------------------------------------

    _ESTADO_ASISTENCIA_LABEL = {
        "P": "Presente",
        "FJ": "Falta Justificada",
        "FI": "Falta Injustificada",
        "R": "Retraso",
        "E": "Excusa",
    }

    # tipo → (título PDF, nombre de hoja Excel)
    _ESTADISTICO_TITULOS = {
        "consolidado_notas": ("Consolidado de Notas", "Consolidado Notas"),
        "consolidado_asistencia": ("Consolidado de Asistencia", "Consolidado Asistencia"),
        "ranking_grupo": ("Ranking del Grupo", "Ranking"),
        "distribucion_desempenos": ("Distribución de Desempeños", "Distribución Desempeños"),
        "estados_asistencia": ("Estados de Asistencia", "Estados Asistencia"),
        "comparativo_periodos": ("Comparativo por Periodos", "Comparativo Periodos"),
        "promedios_area": ("Promedios por Área", "Promedios por Área"),
        "tendencia_asistencia": ("Tendencia de Asistencia", "Tendencia Asistencia"),
    }

    def _filas_estadistico(self, tipo: str, datos) -> list[dict]:
        """Normaliza `datos` (de cualquier tipo) a list[dict] lista para exportar."""
        if tipo == "estados_asistencia":
            return [
                {"Estado": self._ESTADO_ASISTENCIA_LABEL.get(k, k), "Registros": v}
                for k, v in (datos or {}).items()
            ]
        if tipo == "distribucion_desempenos":
            return [{"Nivel de Desempeño": k, "Estudiantes": v} for k, v in (datos or {}).items()]
        if tipo == "comparativo_periodos":
            raw = sanitizar_datos_exportacion(datos if isinstance(datos, list) else [])
            return [
                {
                    "Periodo": r.get("periodo_nombre", r.get("Periodo", "")),
                    "Promedio": r.get("promedio", r.get("Promedio", 0)),
                }
                for r in raw
            ]
        if tipo == "tendencia_asistencia":
            return [
                {"Semana": r.get("semana", ""), "% Asistencia": r.get("porcentaje", 0)}
                for r in (datos if isinstance(datos, list) else [])
            ]
        # Tabulares directos (consolidado_notas/asistencia, ranking, promedios_area, …)
        return sanitizar_datos_exportacion(datos if isinstance(datos, list) else [])

    @staticmethod
    def _inyectar_meta_html(html_str: str, contexto: dict | None) -> str:
        """Inyecta <meta> de grupo/periodo/asignatura para el membrete del PDF."""
        contexto = contexto or {}
        metas = (
            f'<meta name="report-grupo" content="{contexto.get("grupo_nombre", "")}">'
            f'<meta name="report-periodo" content="{contexto.get("periodo_nombre", "")}">'
            f'<meta name="report-asignatura" content="{contexto.get("asignatura_nombre", "")}">'
        )
        return html_str.replace("</head>", f"{metas}</head>", 1)

    def exportar_estadistico(
        self,
        tipo: str,
        datos,
        formato: FormatoInforme | str,
        contexto: dict | None = None,
    ) -> bytes:
        """Exporta un estadístico a Excel o PDF encapsulando todo el pipeline
        (sanitizar + normalizar filas + to_html + inyectar meta + exporter).

        `contexto` puede traer: grupo_id, anio_id (para consolidado_anual),
        grupo_nombre/periodo_nombre/asignatura_nombre (membrete PDF).
        """
        fmt = FormatoInforme(formato) if isinstance(formato, str) else formato
        contexto = contexto or {}
        exporter = self._get_exporter_o_lanzar()

        # consolidado_anual delega en su generador (ya sanitiza internamente)
        if tipo == "consolidado_anual":
            return self.generar_consolidado_anual(
                contexto.get("grupo_id"), contexto.get("anio_id"), formato=fmt
            )

        if tipo not in self._ESTADISTICO_TITULOS:
            raise ValueError(f"Tipo de informe no reconocido: {tipo!r}")

        titulo, nombre_hoja = self._ESTADISTICO_TITULOS[tipo]
        filas = self._filas_estadistico(tipo, datos)

        if fmt == FormatoInforme.EXCEL:
            return exporter.exportar_excel(filas, nombre_hoja=nombre_hoja)
        html = self._datos_a_html(filas, titulo=titulo)
        return exporter.exportar_pdf(self._inyectar_meta_html(html, contexto))

    _ENCABEZADOS_ANCHOS = frozenset({
        "Estudiante", "Documento", "Asignatura", "Área", "Concepto",
        "Observaciones", "Estado", "Nivel de Desempeño", "Periodo", "Semana",
    })

    @staticmethod
    def _datos_a_html(datos: list[dict], titulo: str = "Informe") -> str:
        if not datos:
            return (
                f"<html><head><meta charset='utf-8'></head>"
                f"<body><h1>{titulo}</h1><p>No hay datos para mostrar.</p></body></html>"
            )

        encabezados = list(datos[0].keys())
        usar_vertical = len(encabezados) > 5

        def _celda(val) -> str:
            if val is None:
                return "—"
            s = str(val).strip()
            return s if s and s.lower() != "none" else "—"

        filas_html = "".join(
            "<tr>" + "".join(f"<td>{_celda(fila.get(col))}</td>" for col in encabezados) + "</tr>"
            for fila in datos
        )

        if usar_vertical:
            ths = []
            for col in encabezados:
                if col in InformeService._ENCABEZADOS_ANCHOS:
                    ths.append(f"<th>{col}</th>")
                else:
                    ths.append(f'<th class="v">{col}</th>')
            encabezados_html = "".join(ths)
        else:
            encabezados_html = "".join(f"<th>{col}</th>" for col in encabezados)

        css_vertical = (
            "th.v { writing-mode: vertical-rl; transform: rotate(180deg); "
            "white-space: nowrap; height: 120px; font-size: 9px; padding: 4px 2px; "
            "min-width: 28px; }"
        ) if usar_vertical else ""

        return (
            f"<html><head><meta charset='utf-8'>"
            f"<style>"
            f"body {{ font-family: Arial, sans-serif; font-size: 11px; }}"
            f"h1 {{ font-size: 15px; color: #2B3674; margin-bottom: 8px; }}"
            f"table {{ border-collapse: collapse; width: 100%; }}"
            f"th {{ background-color: #2B6CB0; color: white; padding: 6px 8px; text-align: center; }}"
            f"td {{ border: 1px solid #ddd; padding: 4px 8px; text-align: center; }}"
            f"td:first-child {{ text-align: left; }}"
            f"tr:nth-child(even) {{ background-color: #f5f5f5; }}"
            f"{css_vertical}"
            f"</style></head>"
            f"<body>"
            f"<h1>{titulo}</h1>"
            f"<table>"
            f"<thead><tr>{encabezados_html}</tr></thead>"
            f"<tbody>{filas_html}</tbody>"
            f"</table>"
            f"</body></html>"
        )


def merge_pdfs(pdf_list: list[bytes]) -> bytes:
    """
    Une varios PDF (como bytes) en un único documento PDF.

    Usa pypdf.  Si la lista tiene un solo elemento lo devuelve directamente.
    Lanza ValueError si la lista está vacía.
    """
    if not pdf_list:
        raise ValueError("No hay PDFs para fusionar.")
    if len(pdf_list) == 1:
        return pdf_list[0]

    import io

    from pypdf import PdfReader, PdfWriter

    writer = PdfWriter()
    for pdf_bytes in pdf_list:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        for page in reader.pages:
            writer.add_page(page)

    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


def merge_excels(excel_list: list[tuple[str, bytes]]) -> bytes:
    """
    Combina varios Excel (bytes) en un único libro con una hoja por estudiante.

    Args:
        excel_list: lista de tuplas (nombre_hoja, excel_bytes).
                    El nombre_hoja se trunca a 31 caracteres (límite de Excel).

    Lanza ValueError si la lista está vacía.
    """
    if not excel_list:
        raise ValueError("No hay archivos Excel para fusionar.")

    import copy
    import io

    import openpyxl
    from openpyxl.utils import get_column_letter

    wb_dest = openpyxl.Workbook()
    wb_dest.remove(wb_dest.active)

    for nombre_hoja, excel_bytes in excel_list:
        wb_src = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws_src = wb_src.active
        nombre_safe = nombre_hoja[:31]
        ws_dest = wb_dest.create_sheet(title=nombre_safe)

        for mr in ws_src.merged_cells.ranges:
            ws_dest.merge_cells(str(mr))

        for row in ws_src.iter_rows():
            for cell in row:
                dest_cell = ws_dest.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    dest_cell.font = copy.copy(cell.font)
                    dest_cell.fill = copy.copy(cell.fill)
                    dest_cell.border = copy.copy(cell.border)
                    dest_cell.alignment = copy.copy(cell.alignment)
                    dest_cell.number_format = cell.number_format

        for col_letter, dim in ws_src.column_dimensions.items():
            ws_dest.column_dimensions[col_letter].width = dim.width

        for row_num, dim in ws_src.row_dimensions.items():
            ws_dest.row_dimensions[row_num].height = dim.height

    buf = io.BytesIO()
    wb_dest.save(buf)
    return buf.getvalue()


__all__ = [
    "BoletinesGrupoDTO",
    "FormatoInforme",
    "InformeAsistenciaDTO",
    "InformeNotasDTO",
    "InformeService",
    "merge_excels",
    "merge_pdfs",
    "sanitizar_datos_exportacion",
]
